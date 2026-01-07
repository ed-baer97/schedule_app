"""
API маршруты и функции для связи БД и фронтенда
Содержит все маршруты, которые работают с БД и возвращают данные для фронтенда
"""
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app, send_file
import os
from datetime import datetime, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.db_manager import db, school_db_context, create_school_database, clear_school_database
from app.models.system import School
# Для обратной совместимости
school_db = db
from app.models.school import (
    Subject, ClassGroup, Teacher, ClassLoad, TeacherAssignment,
    PermanentSchedule, TemporarySchedule, Shift, ScheduleSettings,
    PromptClassSubject, PromptClassSubjectTeacher,
    AIConversation, AIConversationMessage, SubjectCabinet, Cabinet,
    SUBJECT_CATEGORIES, SUBJECT_CATEGORY_LANGUAGES,
    SUBJECT_CATEGORY_HUMANITIES, SUBJECT_CATEGORY_NATURAL_MATH
)
from app.services.excel_loader import load_class_load_excel, load_teacher_assignments_excel, load_teacher_contacts_excel, load_cabinets_excel
from app.services.telegram_bot import send_schedule_to_all_teachers, send_temporary_changes_to_all_teachers, send_temporary_changes_to_teacher
from app.core.auth import admin_required, get_current_school_id, current_user
import re

def get_class_group(class_name):
    """
    Определяет группу класса: 'primary' (1-4, начальная школа) или 'secondary' (5-11, старшая школа)
    
    Args:
        class_name: Название класса (например, "1А", "5Б", "11В")
    
    Returns:
        str: 'primary' для 1-4 классов, 'secondary' для 5-11 классов, None если не удалось определить
    """
    if not class_name:
        return None
    
    # Извлекаем число из названия класса (например, "1А" -> 1, "11В" -> 11)
    match = re.match(r'^(\d+)', str(class_name).strip())
    if match:
        class_number = int(match.group(1))
        if 1 <= class_number <= 4:
            return 'primary'
        elif 5 <= class_number <= 11:
            return 'secondary'
    
    return None

def sort_classes_key(class_name):
    """
    Функция для правильной сортировки классов по названию.
    Сортирует по числовой части (1, 2, ..., 9, 10, 11), а затем по буквенной части.
    
    Args:
        class_name: Название класса (например, "1А", "10Б", "11В")
    
    Returns:
        tuple: (число, буква) для сортировки
    """
    if not class_name:
        return (999, '')  # Классы без названия в конец
    
    class_name_str = str(class_name).strip()
    
    # Извлекаем число и букву из названия класса
    match = re.match(r'^(\d+)([А-Яа-яA-Za-z]*)', class_name_str)
    if match:
        number = int(match.group(1))
        letter = match.group(2).upper() if match.group(2) else ''
        return (number, letter)
    
    # Если не удалось распарсить, возвращаем как есть (в конец)
    return (999, class_name_str)

def get_sorted_classes(query=None):
    """
    Получает классы из БД и сортирует их правильно (10-11 после 9).
    
    Args:
        query: SQLAlchemy query объект (опционально). Если не указан, получает все классы.
    
    Returns:
        list: Отсортированный список классов
    """
    if query is None:
        classes = db.session.query(ClassGroup).all()
    else:
        classes = query.all()
    
    # Сортируем классы по правильному ключу
    return sorted(classes, key=lambda cls: sort_classes_key(cls.name))

def ensure_ai_tables_exist():
    """Проверяет и создает таблицы для диалога с ИИ, если их нет"""
    try:
        from flask import current_app
        from sqlalchemy import inspect
        
        engine = db.get_engine(current_app, bind='school')
        inspector = inspect(engine)
        existing_tables = inspector.get_table_names()
        
        if 'ai_conversations' not in existing_tables:
            print("🔄 Создание таблиц для диалога с ИИ...")
            AIConversation.__table__.create(engine, checkfirst=True)
            AIConversationMessage.__table__.create(engine, checkfirst=True)
            print("✅ Таблицы для диалога с ИИ созданы")
        
        # Создаем таблицу shift_classes, если её нет
        from app.models.school import ShiftClass
        if 'shift_classes' not in existing_tables:
            print("🔄 Создание таблицы shift_classes...")
            ShiftClass.__table__.create(engine, checkfirst=True)
            print("✅ Таблица shift_classes создана")
    except Exception as e:
        print(f"⚠️ Ошибка при проверке/создании таблиц: {e}")
        import traceback
        traceback.print_exc()

# Создаем Blueprint для API маршрутов
api_bp = Blueprint('api', __name__)

# Регистрируем Blueprint'ы для маршрутов
# ВАЖНО: Регистрируем здесь, чтобы они были доступны при импорте api_bp
from app.routes import admin, teachers, subjects, schedule, telegram, cabinets, loads
api_bp.register_blueprint(admin.admin_bp)
api_bp.register_blueprint(teachers.teachers_bp)
api_bp.register_blueprint(subjects.subjects_bp)
api_bp.register_blueprint(schedule.schedule_bp)
api_bp.register_blueprint(telegram.telegram_bp)
api_bp.register_blueprint(cabinets.cabinets_bp)
api_bp.register_blueprint(loads.loads_bp)

# AI routes removed - only Telegram bot is kept

# ==================== АДМИН ПАНЕЛЬ ====================

@api_bp.route('/admin')
@admin_required
def admin_index():
    """Главная страница админ-панели"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    # Получаем информацию о школе для отображения названия
    school = School.query.get(school_id)
    school_name = school.name if school else ''
    
    # Убеждаемся, что БД школы существует
    db_path = os.path.join(os.path.dirname(__file__), 'databases', f'school_{school_id}.db')
    if not os.path.exists(db_path):
        try:
            create_school_database(school_id)
        except Exception as e:
            flash(f'Ошибка при создании БД школы: {str(e)}', 'danger')
            return redirect(url_for('logout'))
    
    try:
        with school_db_context(school_id):
            # Получаем активную смену
            active_shift = db.session.query(Shift).filter_by(is_active=True).first()
            if not active_shift:
                shifts = db.session.query(Shift).all()
                if shifts:
                    active_shift = shifts[0]
                    active_shift.is_active = True
                    db.session.commit()
                else:
                    active_shift = Shift(name='Первая смена', is_active=True)
                    db.session.add(active_shift)
                    db.session.commit()
            
            # Получаем предметы
            try:
                subjects = db.session.query(Subject).join(ClassLoad).filter(
                    ClassLoad.shift_id == active_shift.id
                ).distinct().order_by(Subject.name).all()
            except Exception:
                subjects = db.session.query(Subject).order_by(Subject.name).all()
            
            return render_template('admin/index.html', subjects=subjects, current_user=current_user, school_name=school_name)
    except Exception as e:
        flash(f'Ошибка при загрузке данных: {str(e)}', 'danger')
        import traceback
        traceback.print_exc()
        return redirect(url_for('logout'))

@api_bp.route('/admin/teachers')
@admin_required
def teachers_list():
    """Список учителей"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        from app.models.school import CabinetTeacher, Cabinet, ClassGroup
        teachers = db.session.query(Teacher).order_by(Teacher.full_name).all()
        classes = get_sorted_classes()
        # Загружаем связи учителей с классами и кабинетами
        for teacher in teachers:
            # Получаем классы учителя через промежуточную таблицу
            from app.models.school import _get_teacher_classes_table
            teacher_classes_table = _get_teacher_classes_table()
            class_ids = db.session.query(teacher_classes_table.c.class_id).filter(
                teacher_classes_table.c.teacher_id == teacher.id
            ).all()
            teacher.classes_list = [row[0] for row in class_ids]
            
            # Получаем кабинеты учителя через CabinetTeacher
            cabinet_teachers = db.session.query(CabinetTeacher).filter_by(teacher_id=teacher.id).all()
            teacher.cabinets_list = []
            for ct in cabinet_teachers:
                cabinet = db.session.query(Cabinet).filter_by(id=ct.cabinet_id).first()
                if cabinet:
                    teacher.cabinets_list.append(cabinet.name)
            
            # Получаем классы с названиями
            if teacher.classes_list:
                teacher_classes_objs = db.session.query(ClassGroup).filter(ClassGroup.id.in_(teacher.classes_list)).all()
                teacher.classes_names = [c.name for c in teacher_classes_objs]
            else:
                teacher.classes_names = []
        return render_template('admin/teachers.html', teachers=teachers, classes=classes)

# Маршрут /admin/subjects перенесен в app/routes/subjects.py для избежания дублирования

@api_bp.route('/admin/matrix/<subject_name>')
@admin_required
def subject_matrix(subject_name):
    """Матрица предметов"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        subject = db.session.query(Subject).filter_by(name=subject_name).first_or_404()
        
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            return redirect(url_for('api.admin_index'))
        
        # Сначала пытаемся получить учителей для активной смены
        teachers = db.session.query(Teacher).join(TeacherAssignment).filter(
            TeacherAssignment.subject_id == subject.id,
            TeacherAssignment.shift_id == active_shift.id
        ).distinct().order_by(Teacher.full_name).all()
        
        # Если учителей нет для активной смены, получаем для любой смены
        if not teachers:
            teachers = db.session.query(Teacher).join(TeacherAssignment).filter(
                TeacherAssignment.subject_id == subject.id
            ).distinct().order_by(Teacher.full_name).all()
        
        # Загружаем классы для каждого учителя из TeacherAssignment для этого предмета
        # Это гарантирует, что данные совпадут со страницей "Классы"
        teachers_with_classes = []
        for teacher in teachers:
            # Сначала пытаемся получить назначения для активной смены
            teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                teacher_id=teacher.id,
                subject_id=subject.id,
                shift_id=active_shift.id
            ).all()
            
            # Если нет назначений для активной смены, получаем для любой смены
            if not teacher_assignments:
                teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                    teacher_id=teacher.id,
                    subject_id=subject.id
                ).all()
            
            # Получаем уникальные классы из назначений
            class_ids = list(set([ta.class_id for ta in teacher_assignments if ta.class_id]))
            classes = get_sorted_classes(db.session.query(ClassGroup).filter(ClassGroup.id.in_(class_ids))) if class_ids else []
            
            teachers_with_classes.append({
                'teacher': teacher,
                'classes': classes
            })
        
        if teachers:
            all_teachers = db.session.query(Teacher).filter(
                ~Teacher.id.in_([t.id for t in teachers])
            ).order_by(Teacher.full_name).all()
        else:
            all_teachers = db.session.query(Teacher).order_by(Teacher.full_name).all()

        return render_template('admin/subject_matrix.html',
                               subject=subject, 
                               teachers_with_classes=teachers_with_classes, 
                               teachers=teachers, 
                               all_teachers=all_teachers, 
                               shift_id=active_shift.id)

@api_bp.route('/admin/upload', methods=['GET', 'POST'])
@admin_required
def upload_files():
    """Загрузка Excel файлов"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        if request.method == 'POST':
            shift_id = request.form.get('shift_id', type=int)
            shift = None
            if shift_id:
                shift = db.session.query(Shift).filter_by(id=shift_id).first()
                if not shift:
                    flash('Выбранная смена не найдена!', 'error')
                    shifts = db.session.query(Shift).order_by(Shift.id).all()
                    return render_template('admin/upload.html', shifts=shifts)
            
            files_uploaded = False
            
            if 'class_load' in request.files and request.files['class_load'].filename:
                f = request.files['class_load']
                path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'class_load.xlsx')
                f.save(path)
                
                # Если shift_id не указан, функция создаст смены автоматически из листов Excel
                created_shifts = load_class_load_excel(path, shift_id, school_id)
                
                if created_shifts:
                    # Были созданы новые смены
                    shifts_list = ', '.join([f'"{name}"' for name in created_shifts.keys()])
                    flash(f'Создано смен: {len(created_shifts)} ({shifts_list}). Нагрузка классов загружена успешно!', 'success')
                elif shift_id and shift:
                    flash(f'Нагрузка классов загружена успешно для смены "{shift.name}"!', 'success')
                else:
                    flash('Нагрузка классов загружена успешно!', 'success')
                files_uploaded = True

            if 'teacher_assign' in request.files and request.files['teacher_assign'].filename:
                f = request.files['teacher_assign']
                path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_assign.xlsx')
                f.save(path)
                if shift_id:
                    load_teacher_assignments_excel(path, shift_id, school_id)
                    flash(f'Назначения учителей загружены успешно для смены "{shift.name}"!', 'success')
                else:
                    flash('Для загрузки назначений учителей необходимо выбрать смену!', 'error')
                files_uploaded = True

            if 'teacher_contacts' in request.files and request.files['teacher_contacts'].filename:
                f = request.files['teacher_contacts']
                path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_contacts.xlsx')
                f.save(path)
                try:
                    updated, created = load_teacher_contacts_excel(path, shift_id, school_id)
                    flash(f'Контакты учителей загружены успешно! Обновлено: {updated}, создано: {created}', 'success')
                except Exception as e:
                    flash(f'Ошибка при загрузке контактов учителей: {str(e)}', 'error')
                files_uploaded = True

            if 'teacher_cabinets' in request.files and request.files['teacher_cabinets'].filename:
                f = request.files['teacher_cabinets']
                path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_cabinets.xlsx')
                f.save(path)
                try:
                    cabinets_created, links_created, skipped = load_cabinets_excel(path, school_id)
                    flash(f'Кабинеты и учителя загружены успешно! Создано кабинетов: {cabinets_created}, связей: {links_created}' + 
                          (f', пропущено учителей (не найдены в БД): {skipped}' if skipped > 0 else ''), 'success')
                except Exception as e:
                    flash(f'Ошибка при загрузке кабинетов: {str(e)}', 'error')
                files_uploaded = True
            
            if not files_uploaded:
                flash('Выберите хотя бы один файл для загрузки', 'warning')

            return redirect(url_for('api.admin_index'))
        
        shifts = db.session.query(Shift).order_by(Shift.id).all()
        if not shifts:
            default_shift = Shift(name='Первая смена', is_active=True)
            db.session.add(default_shift)
            db.session.commit()
            shifts = [default_shift]
        
        return render_template('admin/upload.html', shifts=shifts)


@api_bp.route('/admin/upload/class-load', methods=['POST'])
@admin_required
def upload_class_load_single():
    """Загрузка только файла Часы_Класс_Предмет"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    shift_id = request.form.get('shift_id', type=int)
    
    with school_db_context(school_id):
        try:
            path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'class_load.xlsx')
            f.save(path)
            
            created_shifts = load_class_load_excel(path, shift_id, school_id)
            
            if created_shifts:
                shifts_list = ', '.join([f'"{name}"' for name in created_shifts.keys()])
                message = f'Создано смен: {len(created_shifts)} ({shifts_list}). Нагрузка классов загружена успешно!'
            elif shift_id:
                shift = db.session.query(Shift).filter_by(id=shift_id).first()
                message = f'Нагрузка классов загружена успешно для смены "{shift.name if shift else shift_id}"!'
            else:
                message = 'Нагрузка классов загружена успешно!'
            
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/upload/teacher-assign', methods=['POST'])
@admin_required
def upload_teacher_assign_single():
    """Загрузка только файла Учителя_Предмет"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    shift_id = request.form.get('shift_id', type=int)
    
    if not shift_id:
        return jsonify({'success': False, 'error': 'Для загрузки назначений учителей необходимо выбрать смену!'}), 400
    
    with school_db_context(school_id):
        try:
            path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_assign.xlsx')
            f.save(path)
            
            load_teacher_assignments_excel(path, shift_id, school_id)
            
            shift = db.session.query(Shift).filter_by(id=shift_id).first()
            message = f'Назначения учителей загружены успешно для смены "{shift.name if shift else shift_id}"!'
            
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/upload/teacher-contacts', methods=['POST'])
@admin_required
def upload_teacher_contacts_single():
    """Загрузка только файла Учителя_Контакты"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    shift_id = request.form.get('shift_id', type=int)
    
    with school_db_context(school_id):
        try:
            path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_contacts.xlsx')
            f.save(path)
            
            updated, created = load_teacher_contacts_excel(path, shift_id, school_id)
            message = f'Контакты учителей загружены успешно! Обновлено: {updated}, создано: {created}'
            
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500


@api_bp.route('/admin/upload/teacher-cabinets', methods=['POST'])
@admin_required
def upload_teacher_cabinets_single():
    """Загрузка только файла Учителя_Кабинет"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
    
    with school_db_context(school_id):
        try:
            path = os.path.join(current_app.config['UPLOAD_FOLDER'], 'teacher_cabinets.xlsx')
            f.save(path)
            
            cabinets_created, links_created, skipped = load_cabinets_excel(path, school_id)
            message = f'Кабинеты и учителя загружены успешно! Создано кабинетов: {cabinets_created}, связей: {links_created}'
            if skipped > 0:
                message += f', пропущено учителей (не найдены в БД): {skipped}'
            
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule')
@admin_required
def schedule():
    """Страница расписания"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        # НЕ инициализируем связь Teacher.classes - используем прямые запросы к промежуточной таблице
        # Это позволяет избежать проблем с проверкой внешних ключей при инициализации
        
        # Проверяем и создаем таблицы, если их нет
        ensure_ai_tables_exist()
        
        shifts = db.session.query(Shift).order_by(Shift.id).all()
        if not shifts:
            default_shift = Shift(name='Первая смена', is_active=True)
            db.session.add(default_shift)
            db.session.commit()
            shifts = [default_shift]
        
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            active_shift = shifts[0]
            active_shift.is_active = True
            db.session.commit()
        
        active_shift_id = active_shift.id
        
        # Получаем классы, назначенные активной смене
        from app.models.school import ShiftClass
        assigned_class_ids = set()
        try:
            assigned_class_ids = set(
                sc.class_id for sc in db.session.query(ShiftClass).filter_by(shift_id=active_shift_id).all()
            )
        except Exception as e:
            print(f"⚠️ Ошибка при получении классов смены: {e}")
        
        # Если нет явно назначенных классов, используем все классы (обратная совместимость)
        if not assigned_class_ids:
            print(f"⚠️ Для смены {active_shift_id} нет явно назначенных классов, показываем все классы")
            classes = get_sorted_classes()
        else:
            print(f"✅ Для смены {active_shift_id} найдено {len(assigned_class_ids)} назначенных классов")
            # Фильтруем только назначенные классы
            classes = get_sorted_classes(
                db.session.query(ClassGroup).filter(ClassGroup.id.in_(assigned_class_ids))
            )
        
        subjects = db.session.query(Subject).order_by(Subject.name).all()
        teachers = db.session.query(Teacher).order_by(Teacher.full_name).all()
        
        settings = {}
        schedule_settings = db.session.query(ScheduleSettings).filter_by(shift_id=active_shift_id).all()
        for setting in schedule_settings:
            settings[setting.day_of_week] = setting.lessons_count
        
        if not settings:
            for day in range(1, 8):
                setting = ScheduleSettings(shift_id=active_shift_id, day_of_week=day, lessons_count=6)
                db.session.add(setting)
                settings[day] = 6
            db.session.commit()
        
        # Получаем расписание только для классов этой смены
        if assigned_class_ids:
            permanent_schedule = db.session.query(PermanentSchedule).filter_by(shift_id=active_shift_id).filter(
                PermanentSchedule.class_id.in_(assigned_class_ids)
            ).join(ClassGroup).join(Subject).join(Teacher).order_by(
                PermanentSchedule.day_of_week,
                PermanentSchedule.lesson_number,
                ClassGroup.name
            ).all()
        else:
            # Если нет назначенных классов, показываем все (обратная совместимость)
            permanent_schedule = db.session.query(PermanentSchedule).filter_by(shift_id=active_shift_id).join(
                ClassGroup).join(Subject).join(Teacher).order_by(
                PermanentSchedule.day_of_week,
                PermanentSchedule.lesson_number,
                ClassGroup.name
            ).all()
        
        schedule_data = []
        for item in permanent_schedule:
            schedule_data.append({
                'id': item.id,
                'day_of_week': item.day_of_week,
                'lesson_number': item.lesson_number,
                'class_id': item.class_id,
                'subject_name': item.subject.name,
                'teacher_name': item.teacher.full_name,
                'cabinet': item.cabinet or ''
            })
        
        classes_list = [{'id': cls.id, 'name': cls.name} for cls in classes]
        teachers_list = [{'id': t.id, 'full_name': t.full_name} for t in teachers] if teachers else []
        subjects_list = [{'id': s.id, 'name': s.name} for s in subjects] if subjects else []
        
        return render_template('admin/schedule.html',
                             classes=classes,
                             subjects=subjects,
                             teachers=teachers,
                             teachers_list=teachers_list,
                             subjects_list=subjects_list,
                             shifts=shifts,
                             active_shift_id=active_shift_id,
                             schedule_data=schedule_data,
                             lessons_count=settings,
                             classes_list=classes_list)

@api_bp.route('/admin/clear')
@admin_required
def clear_db():
    """Очистить БД школы"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    if request.args.get('confirm') == 'yes':
        try:
            if clear_school_database(school_id):
                flash('База данных школы полностью очищена!', 'warning')
            else:
                flash('Ошибка при очистке базы данных', 'danger')
        except Exception as e:
            flash(f'Ошибка: {str(e)}', 'danger')
    return redirect(url_for('api.admin_index'))

# ==================== УЧИТЕЛЯ (CRUD) ====================

@api_bp.route('/admin/teachers/create', methods=['POST'])
@admin_required
def create_teacher():
    """Создать учителя"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    full_name = data.get('full_name', '').strip()
    phone = data.get('phone')
    phone = phone.strip() if phone else None
    telegram_id = data.get('telegram_id')
    telegram_id = telegram_id.strip() if telegram_id else None

    if not full_name:
        return jsonify({'success': False, 'error': 'Полное имя обязательно'}), 400

    try:
        with school_db_context(school_id):
            existing = db.session.query(Teacher).filter_by(full_name=full_name).first()
            if existing:
                return jsonify({'success': False, 'error': 'Учитель с таким именем уже существует'}), 400

            name_parts = full_name.split()
            if len(name_parts) >= 2:
                short_name = ".".join([n[0] + "." for n in name_parts[:2]])
            else:
                short_name = full_name[:30]

            teacher = Teacher(
                full_name=full_name,
                short_name=short_name,
                phone=phone,
                telegram_id=telegram_id
            )
            db.session.add(teacher)
            db.session.commit()

            return jsonify({'success': True, 'teacher_id': teacher.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/teachers/update/<int:teacher_id>', methods=['POST'])
@admin_required
def update_teacher(teacher_id):
    """Обновить учителя"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    full_name = data.get('full_name', '').strip()
    phone = data.get('phone')
    phone = phone.strip() if phone else None

    if not full_name:
        return jsonify({'success': False, 'error': 'Полное имя обязательно'}), 400

    try:
        with school_db_context(school_id):
            teacher = db.session.query(Teacher).filter_by(id=teacher_id).first_or_404()
            
            existing = db.session.query(Teacher).filter_by(full_name=full_name).first()
            if existing and existing.id != teacher_id:
                return jsonify({'success': False, 'error': 'Учитель с таким именем уже существует'}), 400

            teacher.full_name = full_name
            teacher.phone = phone
            telegram_id = data.get('telegram_id')
            teacher.telegram_id = telegram_id.strip() if telegram_id else None
            
            name_parts = full_name.split()
            if len(name_parts) >= 2:
                teacher.short_name = ".".join([n[0] + "." for n in name_parts[:2]])
            else:
                teacher.short_name = full_name[:30]
            
            db.session.commit()
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/teachers/delete/<int:teacher_id>', methods=['POST'])
@admin_required
def delete_teacher(teacher_id):
    """Удалить учителя"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    try:
        with school_db_context(school_id):
            teacher = db.session.query(Teacher).filter_by(id=teacher_id).first_or_404()
            
            # Удаляем связи с классами (автоматически через CASCADE, но лучше явно)
            from app.models.school import _get_teacher_classes_table
            teacher_classes = _get_teacher_classes_table()
            db.session.execute(teacher_classes.delete().where(teacher_classes.c.teacher_id == teacher_id))
            
            db.session.query(TeacherAssignment).filter_by(teacher_id=teacher_id).delete()
            db.session.query(PermanentSchedule).filter_by(teacher_id=teacher_id).delete()
            db.session.query(TemporarySchedule).filter_by(teacher_id=teacher_id).delete()
            
            db.session.delete(teacher)
            db.session.commit()
            
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== МАТРИЦА ПРЕДМЕТОВ ====================

@api_bp.route('/admin/update_hours', methods=['POST'])
@admin_required
def update_hours():
    """Обновить количество часов для учителя по предмету"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    hours = data.get('hours', 0)
    
    with school_db_context(school_id):
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            return jsonify({'success': False, 'error': 'Нет активной смены'}), 400
        
        shift_id = active_shift.id

        assignment = db.session.query(TeacherAssignment).filter_by(
            shift_id=shift_id,
            teacher_id=teacher_id, 
            subject_id=subject_id, 
            class_id=class_id
        ).first()

        if assignment:
            assignment.hours_per_week = hours
        else:
            assignment = TeacherAssignment(
                shift_id=shift_id,
                teacher_id=teacher_id,
                subject_id=subject_id,
                class_id=class_id,
                hours_per_week=hours
            )
            db.session.add(assignment)

        db.session.commit()

        assigned = sum(
            ta.hours_per_week for ta in db.session.query(TeacherAssignment).filter_by(
                shift_id=shift_id,
                subject_id=subject_id, 
                class_id=class_id
            ).all()
        )

        load = db.session.query(ClassLoad).filter_by(shift_id=shift_id, class_id=class_id, subject_id=subject_id).first()
        required = load.hours_per_week if load else 0
        diff = required - assigned

        return jsonify({'assigned': assigned, 'diff': diff})

@api_bp.route('/admin/add_teacher_to_subject', methods=['POST'])
@admin_required
def add_teacher_to_subject():
    """Добавить учителя к предмету
    
    ВАЖНО: 
    - Учитель добавляется к предмету БЕЗ классов (классы назначаются отдельно)
    - Учитель может преподавать несколько предметов
    """
    import logging
    logger = logging.getLogger(__name__)
    
    school_id = get_current_school_id()
    if not school_id:
        logger.error("add_teacher_to_subject: Школа не найдена")
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    subject_id = data.get('subject_id')
    
    logger.info(f"add_teacher_to_subject: teacher_id={teacher_id}, subject_id={subject_id}")
    
    if not teacher_id or not subject_id:
        logger.error(f"add_teacher_to_subject: Не указаны teacher_id или subject_id")
        return jsonify({'success': False, 'error': 'Не указаны teacher_id или subject_id'}), 400
    
    with school_db_context(school_id):
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            logger.error("add_teacher_to_subject: Нет активной смены")
            return jsonify({'success': False, 'error': 'Нет активной смены'}), 400
        
        shift_id = active_shift.id
        logger.info(f"add_teacher_to_subject: shift_id={shift_id}")

        # Учитель может преподавать несколько предметов
        # Проверяем только, не добавлен ли уже учитель к этому предмету
        existing_assignment = db.session.query(TeacherAssignment).filter_by(
            shift_id=shift_id,
            teacher_id=teacher_id,
            subject_id=subject_id
        ).first()
        
        if existing_assignment:
            logger.warning(f"add_teacher_to_subject: Учитель уже добавлен к этому предмету")
            return jsonify({'success': False, 'error': 'Учитель уже добавлен к этому предмету'}), 400

        # Учитель добавляется к предмету БЕЗ автоматического назначения всех классов
        # Создаем TeacherAssignment только для одного класса (первого доступного) с hours_per_week=0
        # Это нужно для отображения учителя в списке, но классы будут назначены отдельно
        
        # Получаем первый класс, для которого есть ClassLoad для этого предмета
        first_class = db.session.query(ClassGroup).join(ClassLoad).filter(
            ClassLoad.subject_id == subject_id,
            ClassLoad.shift_id.is_(None)
        ).first()
        
        # Если нет классов с ClassLoad shift_id=None, получаем первый класс с ClassLoad для этого предмета
        if not first_class:
            first_class = db.session.query(ClassGroup).join(ClassLoad).filter(
                ClassLoad.subject_id == subject_id
            ).first()
        
        # Если все еще нет классов, получаем первый класс вообще
        if not first_class:
            first_class = db.session.query(ClassGroup).first()
        
        if not first_class:
            logger.error("add_teacher_to_subject: Нет классов в базе данных")
            return jsonify({'success': False, 'error': 'Нет классов в базе данных'}), 400

        # Создаем TeacherAssignment для одного класса с hours_per_week=0
        # Это маркер того, что учитель добавлен к предмету, но классы еще не назначены
        try:
            assignment = TeacherAssignment(
                shift_id=shift_id,
                teacher_id=teacher_id,
                subject_id=subject_id,
                class_id=first_class.id,
                hours_per_week=0
            )
            db.session.add(assignment)
            db.session.commit()
            logger.info(f"add_teacher_to_subject: Успешно добавлен учитель {teacher_id} к предмету {subject_id}")
            return jsonify({'success': True, 'message': 'Учитель добавлен к предмету. Теперь назначьте классы через кнопку "Классы".'})
        except Exception as e:
            db.session.rollback()
            logger.error(f"add_teacher_to_subject: Ошибка при создании назначения: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return jsonify({'success': False, 'error': f'Ошибка при добавлении учителя: {str(e)}'}), 500

@api_bp.route('/admin/remove_teacher_from_subject', methods=['POST'])
@admin_required
def remove_teacher_from_subject():
    """Удалить учителя из предмета
    
    ВАЖНО: Учителей без классов можно удалять только вручную через этот интерфейс.
    При загрузке из Excel учителя добавляются без классов, их нужно назначать вручную.
    """
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    subject_id = data.get('subject_id')
    
    with school_db_context(school_id):
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            return jsonify({'success': False, 'error': 'Нет активной смены'}), 400
        
        shift_id = active_shift.id

        try:
            # Получаем все назначения учителя на этот предмет
            assignments = db.session.query(TeacherAssignment).filter_by(
                shift_id=shift_id,
                teacher_id=teacher_id,
                subject_id=subject_id
            ).all()
            
            # Проверяем, есть ли у учителя классы
            has_classes = len(assignments) > 0
            
            # Удаляем все назначения
            for assignment in assignments:
                db.session.delete(assignment)
            
            db.session.commit()
            return jsonify({'success': True, 'message': 'Учитель удален из предмета'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

# ==================== РАСПИСАНИЕ ====================

@api_bp.route('/admin/teachers/<int:teacher_id>/classes', methods=['GET', 'POST'])
@admin_required
def manage_teacher_classes(teacher_id):
    """Управление классами учителя
    
    Если передан subject_id и shift_id, работает с TeacherAssignment для конкретного предмета.
    Иначе работает с общей таблицей teacher_classes.
    """
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    with school_db_context(school_id):
        teacher = db.session.query(Teacher).filter_by(id=teacher_id).first_or_404()
        
        # Проверяем, передан ли subject_id и shift_id (для работы с конкретным предметом)
        subject_id = request.args.get('subject_id', type=int) if request.method == 'GET' else request.get_json().get('subject_id')
        shift_id = request.args.get('shift_id', type=int) if request.method == 'GET' else request.get_json().get('shift_id')
        
        if request.method == 'GET':
            # Получить список классов учителя
            if subject_id and shift_id:
                # Работаем с TeacherAssignment для конкретного предмета
                # Сначала пытаемся получить для активной смены
                teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                    teacher_id=teacher_id,
                    subject_id=subject_id,
                    shift_id=shift_id
                ).all()
                
                # Если нет назначений для активной смены, получаем для любой смены
                if not teacher_assignments:
                    teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                        teacher_id=teacher_id,
                        subject_id=subject_id
                    ).all()
                    # Если есть несколько назначений для разных смен, приоритет отдаем активной смене
                    # Но если их нет для активной смены, берем все
                
                # Если у учителя только одно назначение с hours_per_week = 0,
                # считаем, что классы еще не назначены (это "пустая" запись при добавлении учителя)
                # В этом случае не показываем классы и не ставим галочки
                if len(teacher_assignments) == 1:
                    first_assignment = teacher_assignments[0]
                    hours = getattr(first_assignment, "hours_per_week", None)
                    # Преобразуем в int для надежности
                    try:
                        hours_int = int(hours) if hours is not None else None
                    except (ValueError, TypeError):
                        hours_int = None
                    
                    if hours_int == 0:
                        # Это маркер того, что учитель добавлен к предмету без классов
                        teacher_classes = []
                    else:
                        # Если hours != 0, обрабатываем нормально
                        class_ids_list = [ta.class_id for ta in teacher_assignments if ta.class_id]
                        classes = db.session.query(ClassGroup).filter(ClassGroup.id.in_(class_ids_list)).all() if class_ids_list else []
                        teacher_classes = [{'id': c.id, 'name': c.name} for c in classes]
                elif len(teacher_assignments) == 0:
                    teacher_classes = []
                else:
                    # Если назначений больше одного, обрабатываем нормально
                    class_ids_list = [ta.class_id for ta in teacher_assignments if ta.class_id]
                    classes = db.session.query(ClassGroup).filter(ClassGroup.id.in_(class_ids_list)).all() if class_ids_list else []
                    teacher_classes = [{'id': c.id, 'name': c.name} for c in classes]
            else:
                # Работаем с общей таблицей teacher_classes
                from app.models.school import _get_teacher_classes_table
                teacher_classes_table = _get_teacher_classes_table()
                class_ids = db.session.query(teacher_classes_table.c.class_id).filter(
                    teacher_classes_table.c.teacher_id == teacher_id
                ).all()
                class_ids_list = [row[0] for row in class_ids]
                classes = db.session.query(ClassGroup).filter(ClassGroup.id.in_(class_ids_list)).all()
                teacher_classes = [{'id': c.id, 'name': c.name} for c in classes]
            
            # Получаем все классы для предмета из ClassLoad (общая нагрузка, shift_id = None)
            if subject_id:
                # Получаем классы, для которых есть ClassLoad для этого предмета
                class_loads = db.session.query(ClassLoad).filter_by(
                    subject_id=subject_id,
                    shift_id=None
                ).all()
                
                # Если нет ClassLoad с shift_id=None, получаем все (для обратной совместимости)
                if not class_loads:
                    class_loads = db.session.query(ClassLoad).filter_by(
                        subject_id=subject_id
                    ).all()
                    # Берем только уникальные комбинации (class_id, subject_id)
                    seen = set()
                    unique_loads = []
                    for cl in class_loads:
                        key = (cl.class_id, cl.subject_id)
                        if key not in seen:
                            unique_loads.append(cl)
                            seen.add(key)
                    class_loads = unique_loads
                
                class_ids_from_load = [cl.class_id for cl in class_loads]
                if class_ids_from_load:
                    classes_query = db.session.query(ClassGroup).filter(ClassGroup.id.in_(class_ids_from_load))
                    all_classes = [{'id': c.id, 'name': c.name} for c in get_sorted_classes(classes_query)]
                else:
                    # Если нет ClassLoad для предмета, возвращаем все классы
                    all_classes = [{'id': c.id, 'name': c.name} for c in get_sorted_classes()]
            else:
                # Если subject_id не указан, возвращаем все классы
                all_classes = [{'id': c.id, 'name': c.name} for c in get_sorted_classes()]
            
            return jsonify({
                'success': True,
                'teacher_classes': teacher_classes,
                'all_classes': all_classes
            })
        
        elif request.method == 'POST':
            # Обновить список классов учителя
            data = request.get_json()
            class_ids = data.get('class_ids', [])
            subject_id = data.get('subject_id')
            shift_id = data.get('shift_id')
            
            try:
                if subject_id and shift_id:
                    # Работаем с TeacherAssignment для конкретного предмета
                    # Получаем активную смену, если shift_id не передан или равен 0
                    if not shift_id or shift_id == 0:
                        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
                        if not active_shift:
                            return jsonify({'success': False, 'error': 'Активная смена не найдена'}), 400
                        shift_id = active_shift.id
                    
                    # Удаляем ВСЕ старые TeacherAssignment для этого учителя, предмета и смены
                    deleted_count = db.session.query(TeacherAssignment).filter_by(
                        teacher_id=teacher_id,
                        subject_id=subject_id,
                        shift_id=shift_id
                    ).delete()
                    
                    # Добавляем новые TeacherAssignment для выбранных классов
                    # Если class_ids пустой, создаем одно назначение с hours_per_week=0 как маркер,
                    # что учитель добавлен к предмету, но классы еще не назначены
                    if class_ids:
                        for class_id in class_ids:
                            # Проверяем, есть ли ClassLoad для этого класса и предмета (общая нагрузка, shift_id = None)
                            class_load = db.session.query(ClassLoad).filter_by(
                                class_id=class_id,
                                subject_id=subject_id,
                                shift_id=None
                            ).first()
                            
                            # Если нет ClassLoad с shift_id=None, проверяем любую (для обратной совместимости)
                            if not class_load:
                                class_load = db.session.query(ClassLoad).filter_by(
                                    class_id=class_id,
                                    subject_id=subject_id
                                ).first()
                            
                            if class_load:
                                # Проверяем, нет ли уже такого назначения
                                existing = db.session.query(TeacherAssignment).filter_by(
                                    teacher_id=teacher_id,
                                    subject_id=subject_id,
                                    class_id=class_id,
                                    shift_id=shift_id
                                ).first()
                                
                                if not existing:
                                    # Создаем новое назначение с 0 часами (часы можно будет установить позже)
                                    assignment = TeacherAssignment(
                                        teacher_id=teacher_id,
                                        subject_id=subject_id,
                                        class_id=class_id,
                                        shift_id=shift_id,
                                        hours_per_week=0,
                                        default_cabinet=None
                                    )
                                    db.session.add(assignment)
                    else:
                        # Если class_ids пустой, создаем одно назначение с hours_per_week=0
                        # Это маркер того, что учитель добавлен к предмету, но классы еще не назначены
                        # Получаем первый класс для этого предмета (любой, нужен только для создания записи)
                        first_class_load = db.session.query(ClassLoad).filter_by(
                            subject_id=subject_id,
                            shift_id=None
                        ).first()
                        
                        # Если нет ClassLoad с shift_id=None, получаем любую
                        if not first_class_load:
                            first_class_load = db.session.query(ClassLoad).filter_by(
                                subject_id=subject_id
                            ).first()
                        
                        # Если все еще нет, получаем первый класс вообще
                        if not first_class_load:
                            first_class = db.session.query(ClassGroup).first()
                        else:
                            first_class = db.session.query(ClassGroup).filter_by(id=first_class_load.class_id).first()
                        
                        if first_class:
                            # Проверяем, нет ли уже такого назначения
                            existing = db.session.query(TeacherAssignment).filter_by(
                                teacher_id=teacher_id,
                                subject_id=subject_id,
                                class_id=first_class.id,
                                shift_id=shift_id
                            ).first()
                            
                            if not existing:
                                # Создаем маркерное назначение с hours_per_week=0
                                assignment = TeacherAssignment(
                                    teacher_id=teacher_id,
                                    subject_id=subject_id,
                                    class_id=first_class.id,
                                    shift_id=shift_id,
                                    hours_per_week=0,
                                    default_cabinet=None
                                )
                                db.session.add(assignment)
                else:
                    # Работаем с общей таблицей teacher_classes
                    from app.models.school import _get_teacher_classes_table
                    teacher_classes_table = _get_teacher_classes_table()
                    db.session.execute(
                        teacher_classes_table.delete().where(
                            teacher_classes_table.c.teacher_id == teacher_id
                        )
                    )
                    
                    # Добавляем новые связи
                    if class_ids:
                        for class_id in class_ids:
                            db.session.execute(
                                teacher_classes_table.insert().values(
                                    teacher_id=teacher_id,
                                    class_id=class_id
                                )
                            )
                
                db.session.commit()
                return jsonify({'success': True, 'message': 'Классы учителя обновлены'})
            except Exception as e:
                db.session.rollback()
                import traceback
                traceback.print_exc()
                return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule/teachers/<int:subject_id>')
@admin_required
def get_teachers_for_subject(subject_id):
    """Получить список учителей для предмета (фильтруется по классу для постоянного расписания)"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'teachers': []}), 400
    
    shift_id = request.args.get('shift_id', type=int)
    class_id = request.args.get('class_id', type=int)  # Новый параметр для фильтрации по классу
    
    with school_db_context(school_id):
        if not shift_id:
            active_shift = db.session.query(Shift).filter_by(is_active=True).first()
            if not active_shift:
                return jsonify({'teachers': []})
            shift_id = active_shift.id
        
        # Базовый запрос: учителя, назначенные на предмет в смене
        # Если указан class_id, показываем только учителей, закрепленных за этим классом
        if class_id:
            # Получаем учителей, которые закреплены за этим классом для данного предмета
            # Учитель закреплен за классом, если есть запись в TeacherAssignment с этим class_id
            teacher_ids_for_class = db.session.query(TeacherAssignment.teacher_id).filter(
                TeacherAssignment.subject_id == subject_id,
                TeacherAssignment.shift_id == shift_id,
                TeacherAssignment.class_id == class_id
            ).distinct().all()
            
            teacher_ids_to_include = {row[0] for row in teacher_ids_for_class}
            
            # Если есть учителя, закрепленные за этим классом, показываем их
            if teacher_ids_to_include:
                query = db.session.query(Teacher).filter(Teacher.id.in_(list(teacher_ids_to_include)))
            else:
                # Нет учителей, закрепленных за этим классом для этого предмета - возвращаем пустой список
                query = db.session.query(Teacher).filter(Teacher.id == -1)  # Невозможное условие
        else:
            # Если class_id не указан, возвращаем всех учителей, назначенных на предмет
            query = db.session.query(Teacher).join(TeacherAssignment).filter(
                TeacherAssignment.subject_id == subject_id,
                TeacherAssignment.shift_id == shift_id
            ).distinct()
        
        query = query.distinct()
                    
        # Выполняем запрос
        try:
            teachers = query.order_by(Teacher.full_name).all()
        except Exception as e:
            # Если запрос не удался, возвращаем всех учителей предмета без фильтрации
            print(f"Ошибка при выполнении запроса: {e}")
            import traceback
            traceback.print_exc()
            query = db.session.query(Teacher).join(TeacherAssignment).filter(
                TeacherAssignment.subject_id == subject_id,
                TeacherAssignment.shift_id == shift_id
            ).distinct()
            teachers = query.order_by(Teacher.full_name).all()
        
        # Если не нашли учителей в TeacherAssignment, пробуем получить из PromptClassSubject
        if not teachers and shift_id:
            try:
                # Если указан class_id, ищем учителей для конкретного класса и предмета
                if class_id:
                    pcs = db.session.query(PromptClassSubject).filter_by(
                        shift_id=shift_id,
                        class_id=class_id,
                        subject_id=subject_id
                    ).first()
                    
                    if pcs:
                        # Получаем учителей из PromptClassSubjectTeacher
                        teacher_ids = [pcs_teacher.teacher_id for pcs_teacher in pcs.teachers]
                        if teacher_ids:
                            teachers = db.session.query(Teacher).filter(
                                Teacher.id.in_(teacher_ids)
                            ).order_by(Teacher.full_name).all()
                else:
                    # Если class_id не указан, получаем всех учителей предмета из PromptClassSubject
                    pcs_list = db.session.query(PromptClassSubject).filter_by(
                        shift_id=shift_id,
                        subject_id=subject_id
                    ).all()
                    
                    if pcs_list:
                        # Собираем все уникальные teacher_id
                        teacher_ids = set()
                        for pcs in pcs_list:
                            for pcs_teacher in pcs.teachers:
                                teacher_ids.add(pcs_teacher.teacher_id)
                        
                        if teacher_ids:
                            teachers = db.session.query(Teacher).filter(
                                Teacher.id.in_(list(teacher_ids))
                            ).order_by(Teacher.full_name).all()
            except Exception as e:
                # Если не удалось получить из PromptClassSubject, просто продолжаем с пустым списком
                print(f"Ошибка при получении учителей из PromptClassSubject: {e}")
                import traceback
                traceback.print_exc()
        
        teachers_list = [{'id': t.id, 'name': t.full_name} for t in teachers]
        return jsonify({'teachers': teachers_list})

@api_bp.route('/admin/schedule/cabinets/available')
@admin_required
def get_available_cabinets():
    """Получить список доступных кабинетов для урока с учетом занятости и ограничений"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'cabinets': []}), 400
    
    shift_id = request.args.get('shift_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    teacher_id = request.args.get('teacher_id', type=int)
    class_id = request.args.get('class_id', type=int)
    day_of_week = request.args.get('day_of_week', type=int)
    lesson_number = request.args.get('lesson_number', type=int)
    
    if not all([shift_id, subject_id, teacher_id, class_id, day_of_week, lesson_number]):
        return jsonify({'cabinets': []}), 400
    
    with school_db_context(school_id):
        # Определяем, делится ли предмет на подгруппы
        has_subgroups = False
        prompt_class_subject = db.session.query(PromptClassSubject).filter_by(
            shift_id=shift_id,
            class_id=class_id,
            subject_id=subject_id
        ).first()
        
        if prompt_class_subject:
            has_subgroups = prompt_class_subject.has_subgroups
        else:
            # Проверяем количество учителей для этого класса и предмета
            teachers_count = db.session.query(PromptClassSubjectTeacher).join(
                PromptClassSubject
            ).filter(
                PromptClassSubject.shift_id == shift_id,
                PromptClassSubject.class_id == class_id,
                PromptClassSubject.subject_id == subject_id
            ).count()
            
            if teachers_count == 0:
                # Пробуем через TeacherAssignment
                teachers_count = db.session.query(TeacherAssignment).filter_by(
                    shift_id=shift_id,
                    class_id=class_id,
                    subject_id=subject_id
                ).count()
            
            has_subgroups = teachers_count >= 2
        
        # Получаем доступные кабинеты для учителя используя функцию из schedule_solver_adapter
        from app.services.schedule_solver_adapter import get_available_cabinets_for_teacher
        
        # Получаем кабинет по умолчанию из TeacherAssignment
        default_cabinet = ''
        teacher_assignment = db.session.query(TeacherAssignment).filter_by(
            teacher_id=teacher_id,
            subject_id=subject_id,
            class_id=class_id,
            shift_id=shift_id
        ).first()
        
        if teacher_assignment:
            default_cabinet = teacher_assignment.default_cabinet or ''
        
        available_cabinets = get_available_cabinets_for_teacher(
            teacher_id=teacher_id,
            subject_id=subject_id,
            default_cabinet=default_cabinet,
            has_subgroups=has_subgroups
        )
        
        # Получаем занятые кабинеты в этот день/урок
        occupied_cabinets = set()
        existing_lessons = db.session.query(PermanentSchedule).filter_by(
            shift_id=shift_id,
            day_of_week=day_of_week,
            lesson_number=lesson_number
        ).all()
        
        for lesson in existing_lessons:
            # Если это не подгруппы, кабинет занят полностью
            if not has_subgroups:
                occupied_cabinets.add(lesson.cabinet)
            else:
                # Для подгрупп: кабинет занят, если он уже используется в этой же подгруппе
                # (тот же класс и предмет)
                if lesson.class_id == class_id and lesson.subject_id == subject_id:
                    occupied_cabinets.add(lesson.cabinet)
                else:
                    # Проверяем max_classes_simultaneously для кабинета
                    cabinet_obj = db.session.query(Cabinet).filter_by(name=lesson.cabinet).first()
                    if cabinet_obj:
                        max_classes = cabinet_obj.max_classes_simultaneously or 1
                        # Подсчитываем количество уникальных классов в этом кабинете в этот момент
                        from sqlalchemy import func
                        classes_in_cabinet_query = db.session.query(
                            func.count(func.distinct(PermanentSchedule.class_id))
                        ).filter_by(
                            shift_id=shift_id,
                            day_of_week=day_of_week,
                            lesson_number=lesson_number,
                            cabinet=lesson.cabinet
                        ).scalar()
                        classes_in_cabinet = classes_in_cabinet_query or 0
                        
                        if classes_in_cabinet >= max_classes:
                            occupied_cabinets.add(lesson.cabinet)
        
        # Фильтруем кабинеты: исключаем занятые
        filtered_cabinets = []
        for cab in available_cabinets:
            if cab['name'] not in occupied_cabinets:
                filtered_cabinets.append({
                    'name': cab['name'],
                    'priority': cab.get('priority', 4)
                })
        
        # Сортируем по приоритету
        filtered_cabinets.sort(key=lambda x: x['priority'])
        
        return jsonify({'cabinets': filtered_cabinets})

@api_bp.route('/admin/schedule/data')
@admin_required
def schedule_data():
    """Получить данные расписания"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'error': 'Школа не найдена'}), 400
    
    shift_id = request.args.get('shift_id', type=int)
    
    with school_db_context(school_id):
        if not shift_id:
            shift = db.session.query(Shift).filter_by(is_active=True).first()
            if not shift:
                return jsonify({'error': 'No active shift'}), 400
            shift_id = shift.id
        
        shift = db.session.query(Shift).filter_by(id=shift_id).first()
        if not shift:
            return jsonify({'error': 'Смена не найдена'}), 400
        
        # Проверяем и создаем таблицы, если их нет
        ensure_ai_tables_exist()
        
        # Получаем классы, назначенные этой смене
        from app.models.school import ShiftClass
        assigned_class_ids = set()
        try:
            assigned_class_ids = set(
                sc.class_id for sc in db.session.query(ShiftClass).filter_by(shift_id=shift_id).all()
            )
        except Exception as e:
            print(f"⚠️ Ошибка при получении классов смены: {e}")
        
        settings = {}
        schedule_settings = db.session.query(ScheduleSettings).filter_by(shift_id=shift_id).all()
        for setting in schedule_settings:
            settings[setting.day_of_week] = setting.lessons_count
        
        # Если настроек нет, создаем их по умолчанию (как в функции schedule())
        if not settings:
            for day in range(1, 8):
                setting = ScheduleSettings(shift_id=shift_id, day_of_week=day, lessons_count=6)
                db.session.add(setting)
                settings[day] = 6
            db.session.commit()
        
        # Получаем расписание только для классов этой смены
        if assigned_class_ids:
            permanent_schedule = db.session.query(PermanentSchedule).filter_by(shift_id=shift_id).filter(
                PermanentSchedule.class_id.in_(assigned_class_ids)
            ).join(ClassGroup).join(Subject).join(Teacher).order_by(
                PermanentSchedule.day_of_week,
                PermanentSchedule.lesson_number,
                ClassGroup.name
            ).all()
        else:
            # Если нет назначенных классов, показываем все (обратная совместимость)
            permanent_schedule = db.session.query(PermanentSchedule).filter_by(shift_id=shift_id).join(
                ClassGroup).join(Subject).join(Teacher).order_by(
                PermanentSchedule.day_of_week,
                PermanentSchedule.lesson_number,
                ClassGroup.name
            ).all()
        
        # Логируем для отладки
        print(f"\n📥 Загрузка расписания для отображения:")
        print(f"   Смена ID: {shift_id}")
        if assigned_class_ids:
            print(f"   Классы смены: {len(assigned_class_ids)} классов")
        else:
            print(f"   Классы смены: все классы (обратная совместимость)")
        print(f"   Всего записей найдено: {len(permanent_schedule)}")
        
        # Группируем по классам для статистики
        classes_in_schedule = {}
        for item in permanent_schedule:
            class_id = item.class_id
            if class_id not in classes_in_schedule:
                classes_in_schedule[class_id] = 0
            classes_in_schedule[class_id] += 1
        
        if classes_in_schedule:
            print(f"   Распределение по классам:")
            for class_id, count in sorted(classes_in_schedule.items()):
                cls = db.session.query(ClassGroup).filter_by(id=class_id).first()
                cls_name = cls.name if cls else f"ID {class_id}"
                print(f"      Класс '{cls_name}' (ID {class_id}): {count} уроков")
        
        schedule_data = []
        for item in permanent_schedule:
            schedule_data.append({
                'id': item.id,
                'day_of_week': item.day_of_week,
                'lesson_number': item.lesson_number,
                'class_id': item.class_id,
                'subject_name': item.subject.name,
                'teacher_name': item.teacher.full_name,
                'cabinet': item.cabinet or ''
            })
        
        # Получаем список классов для отображения (только назначенные смене)
        if assigned_class_ids:
            classes_for_display = get_sorted_classes(
                db.session.query(ClassGroup).filter(ClassGroup.id.in_(assigned_class_ids))
            )
        else:
            # Если нет назначенных классов, показываем все (обратная совместимость)
            classes_for_display = get_sorted_classes()
        
        classes_list = [{'id': cls.id, 'name': cls.name} for cls in classes_for_display]
        
        return jsonify({
            'schedule': schedule_data,
            'lessons_count': settings,
            'classes': classes_list
        })

@api_bp.route('/admin/schedule/permanent/add', methods=['POST'])
@admin_required
def add_permanent_schedule():
    """Добавить урок в постоянное расписание"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    
    cabinet = data.get('cabinet', '').strip()
    if not cabinet:
        return jsonify({'success': False, 'error': 'Кабинет обязателен для заполнения'}), 400
    
    shift_id = data.get('shift_id')
    day_of_week = data.get('day_of_week')
    lesson_number = data.get('lesson_number')
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    teacher_id = data.get('teacher_id')
    
    try:
        with school_db_context(school_id):
            shift = db.session.query(Shift).filter_by(id=shift_id).first()
            if not shift:
                return jsonify({'success': False, 'error': 'Смена не найдена'}), 400
            
            # Проверяем, есть ли у добавляемого предмета подгруппы
            prompt_class_subject = db.session.query(PromptClassSubject).filter_by(
                shift_id=shift_id,
                class_id=class_id,
                subject_id=subject_id
            ).first()
            
            # Если в PromptClassSubject нет записи, проверяем по количеству учителей
            if prompt_class_subject:
                current_subject_has_subgroups = prompt_class_subject.has_subgroups
            else:
                # Проверяем количество учителей для этого класса и предмета
                teachers_count = db.session.query(TeacherAssignment).filter_by(
                    shift_id=shift_id,
                    class_id=class_id,
                    subject_id=subject_id
                ).count()
                current_subject_has_subgroups = teachers_count >= 2
            
            # Проверяем, нет ли конфликта между предметами с подгруппами и без подгрупп в одной ячейке
            # Нельзя добавлять предмет с подгруппами, если в ячейке уже есть предмет без подгрупп
            # И наоборот: нельзя добавлять предмет без подгрупп, если в ячейке уже есть предмет с подгруппами
            # Или два предмета без подгрупп в одной ячейке
            existing_lessons_in_cell = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                class_id=class_id
            ).all()
            
            # Проверяем каждый существующий урок
            for existing_lesson in existing_lessons_in_cell:
                # Пропускаем уроки того же предмета (это может быть подгруппа)
                if existing_lesson.subject_id == subject_id:
                    continue
                
                # Проверяем, есть ли у существующего предмета подгруппы
                existing_prompt = db.session.query(PromptClassSubject).filter_by(
                    shift_id=shift_id,
                    class_id=class_id,
                    subject_id=existing_lesson.subject_id
                ).first()
                
                if existing_prompt:
                    existing_subject_has_subgroups = existing_prompt.has_subgroups
                else:
                    # Проверяем по количеству учителей
                    existing_teachers_count = db.session.query(TeacherAssignment).filter_by(
                        shift_id=shift_id,
                        class_id=class_id,
                        subject_id=existing_lesson.subject_id
                    ).count()
                    existing_subject_has_subgroups = existing_teachers_count >= 2
                
                # Конфликт: нельзя смешивать предметы с подгруппами и без подгрупп
                if current_subject_has_subgroups != existing_subject_has_subgroups:
                    existing_subject = db.session.query(Subject).filter_by(id=existing_lesson.subject_id).first()
                    existing_subject_name = existing_subject.name if existing_subject else f"Предмет ID {existing_lesson.subject_id}"
                    current_subject = db.session.query(Subject).filter_by(id=subject_id).first()
                    current_subject_name = current_subject.name if current_subject else f"Предмет ID {subject_id}"
                    
                    if current_subject_has_subgroups:
                        return jsonify({
                            'success': False,
                            'error': f'Нельзя добавить предмет "{current_subject_name}" с подгруппами в ячейку, где уже есть предмет "{existing_subject_name}" без подгрупп.'
                        }), 400
                    else:
                        return jsonify({
                            'success': False,
                            'error': f'Нельзя добавить предмет "{current_subject_name}" без подгрупп в ячейку, где уже есть предмет "{existing_subject_name}" с подгруппами.'
                        }), 400
                
                # Конфликт: нельзя добавлять два предмета без подгрупп в одну ячейку
                if not current_subject_has_subgroups and not existing_subject_has_subgroups:
                    existing_subject = db.session.query(Subject).filter_by(id=existing_lesson.subject_id).first()
                    existing_subject_name = existing_subject.name if existing_subject else f"Предмет ID {existing_lesson.subject_id}"
                    return jsonify({
                        'success': False,
                        'error': f'Нельзя добавить два предмета без подгрупп в одну ячейку. В этой ячейке уже есть предмет "{existing_subject_name}" без подгрупп.'
                    }), 400
            
            # Сначала проверяем, есть ли уже урок по этому предмету в этом классе (подгруппы)
            existing_subgroup_lesson = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                class_id=class_id,
                subject_id=subject_id
            ).first()
            
            is_subgroup = existing_subgroup_lesson is not None
            
            # Проверка: учитель не может вести урок в двух разных классах одновременно
            existing_teacher_lesson = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                teacher_id=teacher_id
            ).first()
            
            if existing_teacher_lesson:
                existing_class = db.session.query(ClassGroup).filter_by(id=existing_teacher_lesson.class_id).first()
                
                # Если это подгруппы (тот же класс и предмет) - разрешаем другого учителя
                # НО только если это другой учитель (не тот же)
                if is_subgroup and existing_teacher_lesson.subject_id == subject_id:
                    # Это подгруппы - разрешаем другого учителя (проверка кабинета будет ниже)
                    # Но если это тот же учитель - это дубликат
                    if existing_teacher_lesson.teacher_id == teacher_id:
                        if existing_teacher_lesson.cabinet == cabinet:
                            return jsonify({
                                'success': False, 
                                'error': 'Этот урок уже добавлен'
                            }), 400
                    # Если это другой учитель в подгруппе - разрешаем (проверка кабинета будет ниже)
                    # Пропускаем остальные проверки для подгрупп
                else:
                    # Это не подгруппы - проверяем обычные конфликты
                    # Если это не тот же класс - это конфликт
                    if existing_teacher_lesson.class_id != class_id:
                        return jsonify({
                            'success': False, 
                            'error': f'Учитель уже ведет урок в классе {existing_class.name} в это время'
                        }), 400
                    
                    # Если это тот же класс, предмет, учитель и кабинет - это дубликат
                    if (existing_teacher_lesson.subject_id == subject_id and 
                        existing_teacher_lesson.cabinet == cabinet):
                        return jsonify({
                            'success': False, 
                            'error': 'Этот урок уже добавлен'
                        }), 400
                    
                    # Если это тот же класс, но другой предмет - это конфликт (учитель не может вести два предмета одновременно)
                    if existing_teacher_lesson.subject_id != subject_id:
                        existing_subject = db.session.query(Subject).filter_by(id=existing_teacher_lesson.subject_id).first()
                        return jsonify({
                            'success': False, 
                            'error': f'Учитель уже ведет {existing_subject.name} в классе {existing_class.name} в это время'
                        }), 400
            
            if is_subgroup:
                # Это подгруппы - проверяем, что кабинет другой
                if existing_subgroup_lesson.cabinet == cabinet:
                    existing_teacher = db.session.query(Teacher).filter_by(id=existing_subgroup_lesson.teacher_id).first()
                    return jsonify({
                        'success': False, 
                        'error': f'Кабинет {cabinet} уже занят учителем {existing_teacher.full_name} в этой подгруппе. Используйте другой кабинет.'
                    }), 400
                # Если кабинет другой - разрешаем (это подгруппы с разными кабинетами)
                # Пропускаем дальнейшие проверки кабинета для подгрупп
                # Также проверяем, что это другой учитель (не тот же)
                if existing_subgroup_lesson.teacher_id == teacher_id:
                    return jsonify({
                        'success': False, 
                        'error': 'Этот учитель уже ведет урок в этой подгруппе. Выберите другого учителя.'
                    }), 400
            else:
                # Это не подгруппы - проверяем, что кабинет не занят
                existing_cabinet_lesson = db.session.query(PermanentSchedule).filter_by(
                    shift_id=shift_id,
                    day_of_week=day_of_week,
                    lesson_number=lesson_number,
                    cabinet=cabinet
                ).first()
                
                if existing_cabinet_lesson and existing_cabinet_lesson.teacher_id != teacher_id:
                    # Кабинет занят другим учителем (не подгруппы)
                    existing_teacher = db.session.query(Teacher).filter_by(id=existing_cabinet_lesson.teacher_id).first()
                    return jsonify({
                        'success': False, 
                        'error': f'Кабинет {cabinet} уже занят учителем {existing_teacher.full_name} в это время'
                    }), 400
            
            class_load = db.session.query(ClassLoad).filter_by(
                shift_id=shift_id,
                class_id=class_id,
                subject_id=subject_id
            ).first()
            
            if not class_load:
                return jsonify({
                    'success': False, 
                    'error': 'Для этого класса не задана нагрузка по данному предмету'
                }), 400
        
            required_hours = class_load.hours_per_week
            
            existing_lessons = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id,
                class_id=class_id,
                subject_id=subject_id
            ).all()
            
            lessons_by_cabinet = {}
            for lesson in existing_lessons:
                cab = lesson.cabinet or 'default'
                if cab not in lessons_by_cabinet:
                    lessons_by_cabinet[cab] = 0
                lessons_by_cabinet[cab] += 1
            
            current_cabinet_lessons = lessons_by_cabinet.get(cabinet, 0)
            if current_cabinet_lessons >= required_hours:
                return jsonify({
                    'success': False, 
                    'error': f'Для подгруппы в кабинете {cabinet} уже добавлено максимальное количество уроков ({required_hours} ч/нед)'
                }), 400
            
            schedule_item = PermanentSchedule(
                shift_id=shift_id,
                day_of_week=day_of_week,
                lesson_number=lesson_number,
                class_id=class_id,
                subject_id=subject_id,
                teacher_id=teacher_id,
                cabinet=cabinet
            )
            db.session.add(schedule_item)
            db.session.commit()
            
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@api_bp.route('/admin/schedule/permanent/delete/<int:schedule_id>', methods=['POST'])
@admin_required
def delete_permanent_schedule(schedule_id):
    """Удалить урок из постоянного расписания"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    try:
        with school_db_context(school_id):
            schedule_item = db.session.query(PermanentSchedule).filter_by(id=schedule_id).first_or_404()
            db.session.delete(schedule_item)
            db.session.commit()
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule/permanent/clear', methods=['POST'])
@admin_required
def clear_permanent_schedule():
    """Очистить все постоянное расписание для текущей смены"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    shift_id = data.get('shift_id')
    
    if not shift_id:
        return jsonify({'success': False, 'error': 'Не указана смена'}), 400
    
    try:
        with school_db_context(school_id):
            # Удаляем все записи постоянного расписания для указанной смены
            deleted_count = db.session.query(PermanentSchedule).filter_by(shift_id=shift_id).delete()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Удалено {deleted_count} уроков из расписания',
                'deleted_count': deleted_count
            })
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule/temporary/add', methods=['POST'])
@admin_required
def add_temporary_schedule():
    """Добавить урок во временное расписание"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    
    try:
        schedule_date = datetime.strptime(data.get('date'), '%Y-%m-%d').date()
        lesson_number = data.get('lesson_number')
        class_id = data.get('class_id')
        subject_id = data.get('subject_id')
        teacher_id = data.get('teacher_id')
        cabinet = data.get('cabinet', '').strip() or None
        
        # Убеждаемся, что работаем в контексте БД школы
        with school_db_context(school_id):
            class_group = db.session.query(ClassGroup).filter_by(id=class_id).first()
            subject = db.session.query(Subject).filter_by(id=subject_id).first()
            teacher = db.session.query(Teacher).filter_by(id=teacher_id).first()
            
            if not class_group or not subject or not teacher:
                return jsonify({'success': False, 'error': 'Неверные данные'}), 400
            
            # Проверяем, есть ли у добавляемого предмета подгруппы
            # Для временного расписания определяем по количеству учителей
            # Сначала пытаемся найти shift_id через ShiftClass
            from app.models.school import ShiftClass
            shift_class = db.session.query(ShiftClass).filter_by(class_id=class_id).first()
            
            if shift_class:
                # Если нашли смену, проверяем через PromptClassSubject
                prompt_class_subject = db.session.query(PromptClassSubject).filter_by(
                    shift_id=shift_class.shift_id,
                    class_id=class_id,
                    subject_id=subject_id
                ).first()
                
                if prompt_class_subject:
                    current_subject_has_subgroups = prompt_class_subject.has_subgroups
                else:
                    # Проверяем по количеству учителей
                    teachers_count = db.session.query(TeacherAssignment).filter_by(
                        shift_id=shift_class.shift_id,
                        class_id=class_id,
                        subject_id=subject_id
                    ).count()
                    current_subject_has_subgroups = teachers_count >= 2
            else:
                # Если смену не нашли, проверяем по количеству учителей (без shift_id)
                teachers_count = db.session.query(TeacherAssignment).filter_by(
                    class_id=class_id,
                    subject_id=subject_id
                ).count()
                current_subject_has_subgroups = teachers_count >= 2
            
            # Проверяем, нет ли конфликта между предметами с подгруппами и без подгрупп в одной ячейке
            # Нельзя добавлять предмет с подгруппами, если в ячейке уже есть предмет без подгрупп
            # И наоборот: нельзя добавлять предмет без подгрупп, если в ячейке уже есть предмет с подгруппами
            # Или два предмета без подгрупп в одной ячейке
            existing_lessons_in_cell = db.session.query(TemporarySchedule).filter_by(
                date=schedule_date,
                lesson_number=lesson_number,
                class_id=class_id
            ).all()
            
            # Проверяем каждый существующий урок
            for existing_lesson in existing_lessons_in_cell:
                # Пропускаем уроки того же предмета
                if existing_lesson.subject_id == subject_id:
                    continue
                
                # Проверяем, есть ли у существующего предмета подгруппы
                if shift_class:
                    existing_prompt = db.session.query(PromptClassSubject).filter_by(
                        shift_id=shift_class.shift_id,
                        class_id=class_id,
                        subject_id=existing_lesson.subject_id
                    ).first()
                    
                    if existing_prompt:
                        existing_subject_has_subgroups = existing_prompt.has_subgroups
                    else:
                        existing_teachers_count = db.session.query(TeacherAssignment).filter_by(
                            shift_id=shift_class.shift_id,
                            class_id=class_id,
                            subject_id=existing_lesson.subject_id
                        ).count()
                        existing_subject_has_subgroups = existing_teachers_count >= 2
                else:
                    existing_teachers_count = db.session.query(TeacherAssignment).filter_by(
                        class_id=class_id,
                        subject_id=existing_lesson.subject_id
                    ).count()
                    existing_subject_has_subgroups = existing_teachers_count >= 2
                
                # Конфликт: нельзя смешивать предметы с подгруппами и без подгрупп
                if current_subject_has_subgroups != existing_subject_has_subgroups:
                    existing_subject = db.session.query(Subject).filter_by(id=existing_lesson.subject_id).first()
                    existing_subject_name = existing_subject.name if existing_subject else f"Предмет ID {existing_lesson.subject_id}"
                    current_subject = db.session.query(Subject).filter_by(id=subject_id).first()
                    current_subject_name = current_subject.name if current_subject else f"Предмет ID {subject_id}"
                    
                    if current_subject_has_subgroups:
                        return jsonify({
                            'success': False,
                            'error': f'Нельзя добавить предмет "{current_subject_name}" с подгруппами в ячейку, где уже есть предмет "{existing_subject_name}" без подгрупп.'
                        }), 400
                    else:
                        return jsonify({
                            'success': False,
                            'error': f'Нельзя добавить предмет "{current_subject_name}" без подгрупп в ячейку, где уже есть предмет "{existing_subject_name}" с подгруппами.'
                        }), 400
                
                # Конфликт: нельзя добавлять два предмета без подгрупп в одну ячейку
                if not current_subject_has_subgroups and not existing_subject_has_subgroups:
                    existing_subject = db.session.query(Subject).filter_by(id=existing_lesson.subject_id).first()
                    existing_subject_name = existing_subject.name if existing_subject else f"Предмет ID {existing_lesson.subject_id}"
                    return jsonify({
                        'success': False,
                        'error': f'Нельзя добавить два предмета без подгрупп в одну ячейку. В этой ячейке уже есть предмет "{existing_subject_name}" без подгрупп.'
                    }), 400
            
            existing_teacher_lesson = db.session.query(TemporarySchedule).filter_by(
                date=schedule_date,
                lesson_number=lesson_number,
                teacher_id=teacher_id
            ).first()
            
            if existing_teacher_lesson:
                existing_class = db.session.query(ClassGroup).filter_by(id=existing_teacher_lesson.class_id).first()
                return jsonify({
                    'success': False, 
                    'error': f'Учитель уже ведет урок в классе {existing_class.name} в это время'
                }), 400
            
            if cabinet:
                existing_cabinet_lesson = db.session.query(TemporarySchedule).filter_by(
                    date=schedule_date,
                    lesson_number=lesson_number,
                    cabinet=cabinet
                ).first()
                
                if existing_cabinet_lesson and existing_cabinet_lesson.teacher_id != teacher_id:
                    existing_teacher = db.session.query(Teacher).filter_by(id=existing_cabinet_lesson.teacher_id).first()
                    return jsonify({
                        'success': False, 
                        'error': f'В кабинете {cabinet} уже ведет урок {existing_teacher.full_name}'
                    }), 400
            
            cabinet_value = cabinet if cabinet else ''
            schedule_item = TemporarySchedule(
                date=schedule_date,
                lesson_number=lesson_number,
                class_id=class_id,
                subject_id=subject_id,
                teacher_id=teacher_id,
                cabinet=cabinet_value
            )
            db.session.add(schedule_item)
            db.session.commit()
            
            try:
                if teacher.telegram_id:
                    send_temporary_changes_to_teacher(teacher, schedule_date, school_id=school_id)
            except Exception as e:
                print(f"Ошибка при отправке уведомления в Telegram: {str(e)}")
            
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@api_bp.route('/admin/schedule/temporary/delete/<int:schedule_id>', methods=['POST'])
@admin_required
def delete_temporary_schedule(schedule_id):
    """Удалить урок из временного расписания"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    try:
        with school_db_context(school_id):
            schedule_item = db.session.query(TemporarySchedule).filter_by(id=schedule_id).first_or_404()
            db.session.delete(schedule_item)
            db.session.commit()
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule/temporary/latest-date')
@admin_required
def temporary_schedule_latest_date():
    """Получить последнюю дату с временным расписанием"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'error': 'Школа не найдена'}), 400
    
    with school_db_context(school_id):
        latest_schedule = db.session.query(TemporarySchedule).order_by(
            TemporarySchedule.date.desc()
        ).first()
        
        if latest_schedule:
            return jsonify({'date': latest_schedule.date.strftime('%Y-%m-%d'), 'has_schedule': True})
        else:
            return jsonify({'date': None, 'has_schedule': False})

@api_bp.route('/admin/schedule/temporary/data')
@admin_required
def temporary_schedule_data():
    """Получить временное расписание на указанную дату"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'error': 'Школа не найдена'}), 400
    
    date_str = request.args.get('date')
    
    if not date_str:
        return jsonify({'error': 'Date parameter is required'}), 400
    
    try:
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
    
    with school_db_context(school_id):
        temporary_schedule = db.session.query(TemporarySchedule).filter_by(date=schedule_date).join(
            ClassGroup).join(Subject).join(Teacher).order_by(
            TemporarySchedule.lesson_number,
            ClassGroup.name
        ).all()
        
        schedule_data = []
        for item in temporary_schedule:
            schedule_data.append({
                'id': item.id,
                'lesson_number': item.lesson_number,
                'class_id': item.class_id,
                'subject_name': item.subject.name,
                'teacher_name': item.teacher.full_name,
                'cabinet': item.cabinet or ''
            })
        
        return jsonify({'schedule': schedule_data})

@api_bp.route('/admin/schedule/export/excel')
@admin_required
def export_schedule_excel():
    """Экспорт расписания в Excel"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'error': 'Школа не найдена'}), 400
    
    shift_id = request.args.get('shift_id', type=int)
    schedule_type = request.args.get('type', 'permanent')  # 'permanent' или 'temporary'
    date_str = request.args.get('date')  # Для временного расписания
    
    with school_db_context(school_id):
        # Получаем информацию о смене
        if not shift_id:
            shift = db.session.query(Shift).filter_by(is_active=True).first()
            if not shift:
                return jsonify({'error': 'Смена не найдена'}), 400
            shift_id = shift.id
        
        shift = db.session.query(Shift).filter_by(id=shift_id).first()
        if not shift:
            return jsonify({'error': 'Смена не найдена'}), 400
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"
        
        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        if schedule_type == 'permanent':
            # Экспорт постоянного расписания в формате таблицы (классы по горизонтали, дни/уроки по вертикали)
            # Получаем настройки количества уроков
            settings = {}
            schedule_settings = db.session.query(ScheduleSettings).filter_by(shift_id=shift_id).all()
            for setting in schedule_settings:
                settings[setting.day_of_week] = setting.lessons_count
            
            # Названия дней недели
            days_names = {
                1: 'Пн',
                2: 'Вт',
                3: 'Ср',
                4: 'Чт',
                5: 'Пт',
                6: 'Сб',
                7: 'Вс'
            }
            
            # Получаем классы
            classes = get_sorted_classes()
            
            # Получаем все расписание
            all_schedule = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id
            ).join(ClassGroup).join(Subject).join(Teacher).order_by(
                PermanentSchedule.day_of_week,
                PermanentSchedule.lesson_number,
                ClassGroup.name
            ).all()
            
            # Создаем словарь для быстрого доступа: (day, lesson, class_id) -> schedule_item
            schedule_dict = {}
            for item in all_schedule:
                key = (item.day_of_week, item.lesson_number, item.class_id)
                if key not in schedule_dict:
                    schedule_dict[key] = []
                schedule_dict[key].append(item)
            
            # Заголовок
            ws['A1'] = f'Постоянное расписание - {shift.name}'
            ws.merge_cells(f'A1:{get_column_letter(len(classes) + 1)}1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment
            
            # Заголовки: День/Урок | Класс1 | Класс2 | ...
            row = 3
            ws.cell(row=row, column=1, value='День/Урок').fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).alignment = center_alignment
            ws.cell(row=row, column=1).border = border
            
            for col, cls in enumerate(classes, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = cls.name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            row += 1
            
            # Заполняем данные по дням и урокам
            for day in range(1, 8):
                if day not in settings or settings[day] == 0:
                    continue
                
                max_lessons = settings[day]
                for lesson_num in range(1, max_lessons + 1):
                    # Ячейка с днем/уроком
                    day_lesson_label = f"{days_names[day]}/{lesson_num}"
                    ws.cell(row=row, column=1, value=day_lesson_label).border = border
                    ws.cell(row=row, column=1).alignment = center_alignment
                    
                    # Заполняем ячейки для каждого класса
                    for col, cls in enumerate(classes, start=2):
                        key = (day, lesson_num, cls.id)
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                        
                        if key in schedule_dict:
                            items = schedule_dict[key]
                            # Если несколько подгрупп, объединяем их
                            cell_lines = []
                            for item in items:
                                line = f"{item.subject.name}\n{item.teacher.full_name}"
                                if item.cabinet:
                                    line += f"\n{item.cabinet}"
                                cell_lines.append(line)
                            cell.value = '\n\n'.join(cell_lines)
                        else:
                            cell.value = ''
                    
                    row += 1
                
                # Пустая строка между днями
                row += 1
            
            # Настраиваем ширину столбцов
            ws.column_dimensions['A'].width = 12
            for col in range(2, len(classes) + 2):
                ws.column_dimensions[get_column_letter(col)].width = 20
                # Устанавливаем высоту строк для многострочного текста
                for r in range(4, row):
                    ws.row_dimensions[r].height = 60
            
            filename = f'Расписание_{shift.name}.xlsx'
        
        else:
            # Экспорт временного расписания
            if not date_str:
                return jsonify({'error': 'Для временного расписания требуется параметр date'}), 400
            
            try:
                schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Неверный формат даты'}), 400
            
            # Заголовок
            date_formatted = schedule_date.strftime('%d.%m.%Y')
            ws['A1'] = f'Временное расписание - {shift.name} - {date_formatted}'
            ws.merge_cells('A1:F1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment
            
            # Заголовки столбцов
            row = 3
            headers = ['Урок', 'Класс', 'Предмет', 'Учитель', 'Кабинет']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            row += 1
            
            # Получаем временное расписание
            temporary_schedule = db.session.query(TemporarySchedule).filter_by(
                date=schedule_date
            ).join(ClassGroup).join(Subject).join(Teacher).order_by(
                TemporarySchedule.lesson_number,
                ClassGroup.name
            ).all()
            
            # Заполняем данные
            for item in temporary_schedule:
                ws.cell(row=row, column=1, value=item.lesson_number).border = border
                ws.cell(row=row, column=2, value=item.class_group.name).border = border
                ws.cell(row=row, column=3, value=item.subject.name).border = border
                ws.cell(row=row, column=4, value=item.teacher.full_name).border = border
                ws.cell(row=row, column=5, value=item.cabinet or '').border = border
                row += 1
            
            filename = f'Временное_расписание_{date_formatted}.xlsx'
        
        # Настраиваем ширину столбцов
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

@api_bp.route('/admin/schedule/temporary/copy', methods=['POST'])
@admin_required
def copy_permanent_to_temporary():
    """Копировать постоянное расписание в временное"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    date_str = data.get('date')
    shift_id = data.get('shift_id')
    day_of_week = data.get('day_of_week')
    
    if not date_str or not shift_id or not day_of_week:
        return jsonify({'success': False, 'error': 'Missing required parameters'}), 400
    
    try:
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format'}), 400
    
    try:
        with school_db_context(school_id):
            shift = db.session.query(Shift).filter_by(id=shift_id).first()
            if not shift:
                return jsonify({'success': False, 'error': 'Смена не найдена'}), 400
            
            # Удаляем существующие записи для этой даты
            db.session.query(TemporarySchedule).filter_by(date=schedule_date).delete()
            db.session.flush()
            
            permanent_schedule = db.session.query(PermanentSchedule).filter_by(
                shift_id=shift_id,
                day_of_week=day_of_week
            ).all()
            
            # Словарь для отслеживания уже обработанных комбинаций (date, lesson_number, class_id, cabinet)
            # Ключ: (date, lesson_number, class_id, cabinet), значение: TemporarySchedule объект
            processed_items = {}
            
            for item in permanent_schedule:
                # Нормализуем кабинет: пустые значения заменяем на '-'
                cabinet_value = item.cabinet.strip() if item.cabinet and item.cabinet.strip() else '-'
                
                # Формируем уникальный ключ для проверки
                unique_key = (schedule_date, item.lesson_number, item.class_id, cabinet_value)
                
                # Проверяем, была ли уже обработана такая комбинация
                if unique_key in processed_items:
                    # Если запись с таким ключом уже существует, это означает подгруппу
                    # Уникальное ограничение не позволяет иметь несколько записей с одинаковым
                    # (date, lesson_number, class_id, cabinet), поэтому мы пропускаем дубликаты
                    # В реальности подгруппы должны иметь разные кабинеты
                    continue
                
                # Создаем новую запись
                temporary_item = TemporarySchedule(
                    date=schedule_date,
                    lesson_number=item.lesson_number,
                    class_id=item.class_id,
                    subject_id=item.subject_id,
                    teacher_id=item.teacher_id,
                    cabinet=cabinet_value
                )
                db.session.add(temporary_item)
                processed_items[unique_key] = temporary_item
            
            db.session.commit()
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== СМЕНЫ И НАСТРОЙКИ ====================

@api_bp.route('/admin/shift/<int:shift_id>/classes', methods=['GET'])
@admin_required
def shift_classes(shift_id):
    """Страница управления классами смены"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    with school_db_context(school_id):
        shift = db.session.query(Shift).filter_by(id=shift_id).first()
        if not shift:
            flash('Смена не найдена', 'danger')
            return redirect(url_for('api.schedule'))
        
        # Получаем все смены для переключения
        all_shifts = db.session.query(Shift).order_by(Shift.id).all()
        
        # Получаем все классы
        all_classes = db.session.query(ClassGroup).order_by(ClassGroup.name).all()
        
        # Проверяем и создаем таблицы, если их нет
        ensure_ai_tables_exist()
        
        # Получаем классы, назначенные этой смене
        from app.models.school import ShiftClass
        assigned_class_ids = set(
            sc.class_id for sc in db.session.query(ShiftClass).filter_by(shift_id=shift_id).all()
        )
        
        return render_template('admin/shift_classes.html', 
                             shift=shift,
                             all_shifts=all_shifts,
                             all_classes=all_classes,
                             assigned_class_ids=assigned_class_ids)

@api_bp.route('/admin/shift/<int:shift_id>/classes/assign', methods=['POST'])
@admin_required
def assign_class_to_shift(shift_id):
    """Назначить класс смене"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    if not class_id:
        return jsonify({'success': False, 'error': 'Не указан класс'}), 400
    
    with school_db_context(school_id):
        # Проверяем и создаем таблицы, если их нет
        ensure_ai_tables_exist()
        
        from app.models.school import ShiftClass
        shift = db.session.query(Shift).filter_by(id=shift_id).first()
        if not shift:
            return jsonify({'success': False, 'error': 'Смена не найдена'}), 400
        
        cls = db.session.query(ClassGroup).filter_by(id=class_id).first()
        if not cls:
            return jsonify({'success': False, 'error': 'Класс не найден'}), 400
        
        # Проверяем, не назначен ли уже
        existing = db.session.query(ShiftClass).filter_by(shift_id=shift_id, class_id=class_id).first()
        if existing:
            return jsonify({'success': False, 'error': 'Класс уже назначен этой смене'}), 400
        
        shift_class = ShiftClass(shift_id=shift_id, class_id=class_id)
        db.session.add(shift_class)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Класс {cls.name} назначен смене {shift.name}'})

@api_bp.route('/admin/shift/<int:shift_id>/classes/remove', methods=['POST'])
@admin_required
def remove_class_from_shift(shift_id):
    """Убрать класс из смены"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    class_id = data.get('class_id')
    
    if not class_id:
        return jsonify({'success': False, 'error': 'Не указан класс'}), 400
    
    with school_db_context(school_id):
        # Проверяем и создаем таблицы, если их нет
        ensure_ai_tables_exist()
        
        from app.models.school import ShiftClass
        shift_class = db.session.query(ShiftClass).filter_by(shift_id=shift_id, class_id=class_id).first()
        if not shift_class:
            return jsonify({'success': False, 'error': 'Класс не назначен этой смене'}), 400
        
        cls = db.session.query(ClassGroup).filter_by(id=class_id).first()
        cls_name = cls.name if cls else f"ID {class_id}"
        
        db.session.delete(shift_class)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Класс {cls_name} убран из смены'})

@api_bp.route('/admin/schedule/shift/add', methods=['POST'])
@admin_required
def add_shift():
    """Добавить смену"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    name = data.get('name', '').strip()
    
    if not name:
        return jsonify({'success': False, 'error': 'Название смены обязательно'}), 400
    
    try:
        with school_db_context(school_id):
            existing = db.session.query(Shift).filter_by(name=name).first()
            if existing:
                return jsonify({'success': False, 'error': 'Смена с таким названием уже существует'}), 400
            
            shift = Shift(name=name, is_active=False)
            db.session.add(shift)
            db.session.commit()
            
            for day in range(1, 8):
                setting = ScheduleSettings(shift_id=shift.id, day_of_week=day, lessons_count=6)
                db.session.add(setting)
            db.session.commit()
            
            return jsonify({'success': True, 'shift_id': shift.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/shift/<int:shift_id>/activate', methods=['POST'])
@admin_required
def activate_shift(shift_id):
    """Активировать смену (сделать её активной)"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    try:
        with school_db_context(school_id):
            # Деактивируем все смены
            db.session.query(Shift).update({Shift.is_active: False})
            
            # Активируем выбранную смену
            shift = db.session.query(Shift).filter_by(id=shift_id).first()
            if not shift:
                return jsonify({'success': False, 'error': 'Смена не найдена'}), 404
            
            shift.is_active = True
            db.session.commit()
            
            return jsonify({'success': True, 'message': f'Смена "{shift.name}" активирована'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/schedule/settings/save', methods=['POST'])
@admin_required
def save_schedule_settings():
    """Сохранить настройки расписания"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    shift_id = data.get('shift_id')
    settings = data.get('settings', {})
    
    try:
        with school_db_context(school_id):
            db.session.query(ScheduleSettings).filter_by(shift_id=shift_id).delete()
            
            for day, count in settings.items():
                setting = ScheduleSettings(
                    shift_id=shift_id,
                    day_of_week=int(day),
                    lessons_count=int(count)
                )
                db.session.add(setting)
            
            db.session.commit()
            return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== TELEGRAM ====================

@api_bp.route('/admin/telegram/send-schedule', methods=['POST'])
def send_schedule_telegram():
    """Отправить расписание всем учителям через Telegram"""
    data = request.get_json()
    shift_id = data.get('shift_id') if data else None
    if shift_id:
        try:
            shift_id = int(shift_id)
        except (ValueError, TypeError):
            shift_id = None
    
    try:
        # Убеждаемся, что работаем в контексте БД школы
        school_id = get_current_school_id()
        if not school_id:
            return jsonify({'success': False, 'error': 'Не удалось определить школу'}), 400
        
        with school_db_context(school_id):
            results = send_schedule_to_all_teachers(shift_id, school_id=school_id)
        
        if 'errors' in results and isinstance(results['errors'], list) and results['errors']:
            error_msg = results['errors'][0] if isinstance(results['errors'][0], str) else 'Ошибка при отправке'
        else:
            error_msg = None
        
        return jsonify({
            'success': True,
            'sent': results.get('success', 0),
            'failed': results.get('failed', 0),
            'errors': results.get('errors', []),
            'message': f"Отправлено: {results.get('success', 0)}, Ошибок: {results.get('failed', 0)}"
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Ошибка отправки: {error_msg}'}), 500

@api_bp.route('/admin/telegram/send-temporary', methods=['POST'])
@admin_required
def send_temporary_telegram():
    """Отправить временное расписание через Telegram"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 400
    
    data = request.get_json()
    date_str = data.get('date')
    
    if not date_str:
        return jsonify({'success': False, 'error': 'Date parameter is required'}), 400
    
    try:
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        # Убеждаемся, что работаем в контексте БД школы
        school_id = get_current_school_id()
        if not school_id:
            return jsonify({'success': False, 'error': 'Не удалось определить школу'}), 400
        
        with school_db_context(school_id):
            results = send_temporary_changes_to_all_teachers(schedule_date, school_id=school_id)
        return jsonify({
            'success': True,
            'sent': results['success'],
            'failed': results['failed'],
            'no_changes': results['no_changes'],
            'errors': results['errors']
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback.print_exc()
        return jsonify({'success': False, 'error': error_msg}), 500

# ==================== AI ====================
# Устаревшие AI маршруты удалены
# Актуальные маршруты находятся в app/routes/ai.py:
# - /admin/ai/generate-solver - генерация через алгоритм
# - /admin/ai/chat - упрощенный чат
# - /admin/ai/conversation/<shift_id>/active - получение истории
# - /admin/ai/apply-all-suggestions - применение предложений

# ==================== КАБИНЕТЫ ====================
# Старые маршруты для кабинетов по предметам удалены
# Используются новые маршруты из app/routes/cabinets.py:
# - GET /admin/cabinets - страница управления кабинетами и учителями
# - POST /admin/cabinets/add - добавить новый кабинет
# - POST /admin/cabinets/add-teacher - добавить учителя к кабинету
# - POST /admin/cabinets/remove-teacher - удалить учителя из кабинета

# Устаревшие AI маршруты удалены - используются маршруты из app/routes/ai.py

# ==================== КАБИНЕТЫ ====================
# Старые маршруты для кабинетов по предметам удалены
# Используются новые маршруты из app/routes/cabinets.py:
# - GET /admin/cabinets - страница управления кабинетами и учителями
# - POST /admin/cabinets/add - добавить новый кабинет
# - POST /admin/cabinets/add-teacher - добавить учителя к кабинету
# - POST /admin/cabinets/remove-teacher - удалить учителя из кабинета

# ==================== ОЧИСТКА БД ДЛЯ АДМИНА ШКОЛЫ ====================

@api_bp.route('/admin/clear-database', methods=['POST'])
@admin_required
def admin_clear_database():
    """Очистить все данные школы (доступно только для админа своей школы)"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Не удалось определить школу'}), 400
    
    # Получаем информацию о школе
    school = School.query.get(school_id)
    if not school:
        return jsonify({'success': False, 'error': 'Школа не найдена'}), 404
    
    # Проверяем, что текущий пользователь - админ этой школы
    if current_user.school_id != school_id:
        return jsonify({'success': False, 'error': 'Нет доступа к этой школе'}), 403
    
    data = request.get_json()
    confirm_text = data.get('confirm', '').strip() if data else ''
    
    # Проверка подтверждения
    if confirm_text != school.name:
        return jsonify({
            'success': False, 
            'error': f'Для подтверждения введите название школы: {school.name}'
        }), 400
    
    try:
        # Очищаем БД школы
        if clear_school_database(school_id):
            return jsonify({
                'success': True,
                'message': f'Все данные школы "{school.name}" успешно удалены'
            })
        else:
            return jsonify({'success': False, 'error': 'Ошибка при очистке БД школы'}), 500
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Ошибка при очистке БД школы: {error_trace}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== НАГРУЗКА КЛАССОВ ПО ПРЕДМЕТАМ ====================

@api_bp.route('/admin/classes')
@admin_required
def classes_page():
    """Страница для просмотра классов с предметами, учителями и часами"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        # Получаем активную смену
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            shifts = db.session.query(Shift).all()
            if shifts:
                active_shift = shifts[0]
            else:
                active_shift = Shift(name='Первая смена', is_active=True)
                db.session.add(active_shift)
                db.session.commit()
        
        # Берем данные напрямую из ClassLoad и TeacherAssignment (как на странице "Предметы")
        # Это гарантирует, что данные будут актуальными и совпадут с данными на странице "Предметы"
        classes = get_sorted_classes()
        classes_data = []
        
        for cls in classes:
            # Нагрузка общая для всех смен (shift_id = None)
            class_loads = db.session.query(ClassLoad).filter_by(
                shift_id=None,
                class_id=cls.id
            ).all()
            
            # Если нет нагрузки с shift_id=None, получаем все (для обратной совместимости)
            if not class_loads:
                all_loads = db.session.query(ClassLoad).filter_by(class_id=cls.id).all()
                # Берем только уникальные комбинации (class_id, subject_id)
                seen = set()
                for cl in all_loads:
                    key = (cl.class_id, cl.subject_id)
                    if key not in seen:
                        class_loads.append(cl)
                        seen.add(key)
            
            for class_load in class_loads:
                subject = db.session.query(Subject).filter_by(id=class_load.subject_id).first()
                if not subject:
                    continue
                
                # Получаем учителей для этого класса и предмета из TeacherAssignment
                # Сначала пытаемся получить назначения для активной смены
                teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                    shift_id=active_shift.id,
                    class_id=cls.id,
                    subject_id=class_load.subject_id
                ).all()
                
                # Если нет назначений для активной смены, получаем для любой смены
                if not teacher_assignments:
                    teacher_assignments = db.session.query(TeacherAssignment).filter_by(
                        class_id=cls.id,
                        subject_id=class_load.subject_id
                    ).all()
                    # Если есть несколько назначений для разных смен, приоритет отдаем активной смене
                    # Но если их нет для активной смены, берем все
                
                teachers = []
                for assignment in teacher_assignments:
                    teacher = db.session.query(Teacher).filter_by(id=assignment.teacher_id).first()
                    if teacher:
                        teachers.append({
                            'teacher_id': teacher.id,
                            'teacher_name': teacher.full_name,
                            'hours_per_week': assignment.hours_per_week or 0,
                            'default_cabinet': assignment.default_cabinet or ''
                        })
                
                classes_data.append({
                    'class_id': cls.id,
                    'class_name': cls.name,
                    'subject_id': subject.id,
                    'subject_name': subject.name,
                    'total_hours_per_week': class_load.hours_per_week,
                    'has_subgroups': len(teachers) >= 2,
                    'teachers': teachers
                })
        
        # Группируем данные по классам для удобного отображения
        classes_dict = {}
        for item in classes_data:
            class_name = item['class_name']
            if class_name not in classes_dict:
                classes_dict[class_name] = {
                    'class_id': item['class_id'],
                    'class_name': class_name,
                    'subjects': []
                }
            
            classes_dict[class_name]['subjects'].append({
                'subject_id': item['subject_id'],
                'subject_name': item['subject_name'],
                'total_hours_per_week': item['total_hours_per_week'],
                'has_subgroups': item['has_subgroups'],
                'teachers': item['teachers']
            })
        
        # Сортируем классы и предметы
        classes_list = sorted(classes_dict.values(), key=lambda x: x['class_name'])
        for cls_data in classes_list:
            cls_data['subjects'].sort(key=lambda x: x['subject_name'])
        
        # Группируем классы по начальной (1-4) и старшей (5-11) школе
        primary_classes = []
        secondary_classes = []
        
        for cls_data in classes_list:
            group = get_class_group(cls_data['class_name'])
            if group == 'primary':
                primary_classes.append(cls_data)
            elif group == 'secondary':
                secondary_classes.append(cls_data)
            else:
                # Если не удалось определить группу, добавляем в старшую школу
                secondary_classes.append(cls_data)
        
        return render_template('admin/classes.html',
                             classes_list=classes_list,
                             primary_classes=primary_classes,
                             secondary_classes=secondary_classes,
                             active_shift=active_shift,
                             current_user=current_user)

@api_bp.route('/admin/class-loads')
@admin_required
def class_loads_page():
    """Страница для редактирования нагрузки классов по предметам"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        # Получаем все классы и предметы
        classes = get_sorted_classes()
        subjects = db.session.query(Subject).order_by(Subject.name).all()
        
        # Получаем текущую нагрузку
        # Сначала пытаемся получить записи с shift_id=None (общая нагрузка)
        class_loads = db.session.query(ClassLoad).filter_by(shift_id=None).all()
        
        # Если нет записей с shift_id=None, получаем все записи (для обратной совместимости)
        # При этом берем только уникальные комбинации (class_id, subject_id), чтобы избежать дубликатов
        if not class_loads:
            all_loads = db.session.query(ClassLoad).all()
            # Группируем по (class_id, subject_id) и берем первую запись для каждой комбинации
            seen = set()
            for cl in all_loads:
                key = (cl.class_id, cl.subject_id)
                if key not in seen:
                    class_loads.append(cl)
                    seen.add(key)
        
        # Создаем словарь для быстрого доступа: (class_id, subject_id) -> hours_per_week
        load_dict = {(cl.class_id, cl.subject_id): cl.hours_per_week for cl in class_loads}
        
        return render_template('admin/class_loads.html', 
                             classes=classes, 
                             subjects=subjects, 
                             load_dict=load_dict,
                             current_user=current_user)

@api_bp.route('/admin/class-loads/auto-fill', methods=['POST'])
@admin_required
def auto_fill_class_loads():
    """Автоматически заполнить нагрузку классов на основе назначений учителей"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Не удалось определить школу'}), 400
    
    try:
        with school_db_context(school_id):
            # Получаем все назначения учителей (для всех смен, суммируем)
            assignments = db.session.query(TeacherAssignment).all()
            
            # Группируем по классу и предмету, суммируя часы для всех смен
            load_dict = {}
            for assignment in assignments:
                if assignment.hours_per_week and assignment.hours_per_week > 0:
                    key = (assignment.class_id, assignment.subject_id)
                    if key not in load_dict:
                        load_dict[key] = 0
                    load_dict[key] += assignment.hours_per_week
            
            # Создаем или обновляем ClassLoad (общая нагрузка для всех смен)
            created_count = 0
            updated_count = 0
            
            for (class_id, subject_id), total_hours in load_dict.items():
                # Проверяем существование класса и предмета
                class_group = db.session.query(ClassGroup).get(class_id)
                subject = db.session.query(Subject).get(subject_id)
                
                if not class_group or not subject:
                    continue
                
                # Ищем существующую нагрузку (общая для всех смен, shift_id = NULL)
                class_load = db.session.query(ClassLoad).filter_by(
                    shift_id=None,
                    class_id=class_id,
                    subject_id=subject_id
                ).first()
                
                if class_load:
                    class_load.hours_per_week = total_hours
                    updated_count += 1
                else:
                    class_load = ClassLoad(
                        shift_id=None,  # Нагрузка общая для всех смен
                        class_id=class_id,
                        subject_id=subject_id,
                        hours_per_week=total_hours
                    )
                    db.session.add(class_load)
                    created_count += 1
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Создано: {created_count}, обновлено: {updated_count}',
                'created': created_count,
                'updated': updated_count
            })
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@api_bp.route('/admin/class-loads/update', methods=['POST'])
@admin_required
def update_class_load():
    """Обновить или создать нагрузку класса по предмету"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Не удалось определить школу'}), 400
    
    data = request.get_json()
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    hours_per_week = data.get('hours_per_week', 0)
    
    if not class_id or not subject_id:
        return jsonify({'success': False, 'error': 'Не указаны class_id или subject_id'}), 400
    
    try:
        hours_per_week = int(hours_per_week)
        if hours_per_week < 0:
            return jsonify({'success': False, 'error': 'Количество часов не может быть отрицательным'}), 400
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Некорректное количество часов'}), 400
    
    try:
        with school_db_context(school_id):
            # Проверяем существование класса и предмета
            class_group = db.session.query(ClassGroup).get(class_id)
            subject = db.session.query(Subject).get(subject_id)
            
            if not class_group:
                return jsonify({'success': False, 'error': 'Класс не найден'}), 404
            if not subject:
                return jsonify({'success': False, 'error': 'Предмет не найден'}), 404
            
            # Ищем существующую нагрузку (общая для всех смен, shift_id = NULL)
            class_load = db.session.query(ClassLoad).filter_by(
                shift_id=None,
                class_id=class_id,
                subject_id=subject_id
            ).first()
            
            if hours_per_week == 0:
                # Удаляем нагрузку, если установлено 0
                if class_load:
                    db.session.delete(class_load)
                    db.session.commit()
                return jsonify({'success': True, 'message': 'Нагрузка удалена'})
            else:
                # Обновляем или создаем нагрузку (без привязки к смене)
                if class_load:
                    class_load.hours_per_week = hours_per_week
                else:
                    class_load = ClassLoad(
                        shift_id=None,  # Нагрузка общая для всех смен
                        class_id=class_id,
                        subject_id=subject_id,
                        hours_per_week=hours_per_week
                    )
                    db.session.add(class_load)
                
                db.session.commit()
                return jsonify({
                    'success': True, 
                    'message': 'Нагрузка обновлена',
                    'class_load_id': class_load.id
                })
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== НАГРУЗКА УЧИТЕЛЕЙ ====================

@api_bp.route('/admin/teacher-workload')
@admin_required
def teacher_workload_page():
    """Страница для отображения нагрузки учителей (сколько часов в неделю)"""
    school_id = get_current_school_id()
    if not school_id:
        flash('Ошибка: школа не найдена', 'danger')
        return redirect(url_for('logout'))
    
    with school_db_context(school_id):
        # Получаем активную смену
        active_shift = db.session.query(Shift).filter_by(is_active=True).first()
        if not active_shift:
            shifts = db.session.query(Shift).all()
            if shifts:
                active_shift = shifts[0]
            else:
                active_shift = Shift(name='Первая смена', is_active=True)
                db.session.add(active_shift)
                db.session.commit()
        
        # Получаем всех учителей
        teachers = db.session.query(Teacher).order_by(Teacher.full_name).all()
        
        # Получаем все назначения учителей для активной смены
        assignments = db.session.query(TeacherAssignment).filter_by(shift_id=active_shift.id).all()
        
        # Получаем все предметы и классы для быстрого доступа
        all_subjects = {s.id: s.name for s in db.session.query(Subject).all()}
        all_classes = {c.id: c.name for c in db.session.query(ClassGroup).all()}
        
        # Вычисляем нагрузку для каждого учителя
        teacher_workload = []
        for teacher in teachers:
            teacher_assignments = [a for a in assignments if a.teacher_id == teacher.id]
            
            # Группируем по предметам и классам
            assignments_by_subject = {}
            total_hours = 0
            for assignment in teacher_assignments:
                # Используем словари вместо relationships для избежания проблем с bind
                subject_name = all_subjects.get(assignment.subject_id, f"ID {assignment.subject_id}")
                class_name = all_classes.get(assignment.class_id, f"ID {assignment.class_id}")
                key = f"{subject_name} ({class_name})"
                
                # Если часы в назначении равны 0, берем часы из ClassLoad
                hours = assignment.hours_per_week
                if hours == 0:
                    class_load = db.session.query(ClassLoad).filter_by(
                        shift_id=active_shift.id,
                        class_id=assignment.class_id,
                        subject_id=assignment.subject_id
                    ).first()
                    if class_load:
                        hours = class_load.hours_per_week
                
                if key not in assignments_by_subject:
                    assignments_by_subject[key] = {
                        'subject': subject_name,
                        'class': class_name,
                        'hours': 0
                    }
                assignments_by_subject[key]['hours'] += hours
                total_hours += hours
            
            teacher_workload.append({
                'teacher': teacher,
                'total_hours': total_hours,
                'assignments': list(assignments_by_subject.values())
            })
        
        # Сортируем по общему количеству часов (по убыванию)
        teacher_workload.sort(key=lambda x: x['total_hours'], reverse=True)
        
        # Вычисляем статистику
        total_hours_all = sum(item['total_hours'] for item in teacher_workload)
        avg_hours = total_hours_all / len(teacher_workload) if teacher_workload else 0
        max_hours = teacher_workload[0]['total_hours'] if teacher_workload else 0
        min_hours = teacher_workload[-1]['total_hours'] if teacher_workload else 0
        teachers_high_load = sum(1 for item in teacher_workload if item['total_hours'] >= 18)
        
        return render_template('admin/teacher_workload.html',
                             teacher_workload=teacher_workload,
                             active_shift=active_shift,
                             current_user=current_user,
                             total_hours_all=total_hours_all,
                             avg_hours=round(avg_hours, 1),
                             max_hours=max_hours,
                             min_hours=min_hours,
                             teachers_high_load=teachers_high_load)

# Регистрация blueprint'ов перенесена в начало файла (после создания api_bp)

